import requests
import time
import openpyxl
from openpyxl import Workbook
import os
import re
from concurrent.futures import ThreadPoolExecutor

API_BASE_URL = "https://api.unpaywall.org/v2/search"
EMAIL = "xlct43@durham.ac.uk"

# Retry decorator with exponential backoff
def retry_request(url, params, retries=3, wait_s=2.0):
    for attempt in range(retries):
        try:
            response = requests.get(url, params=params, timeout=30)
            response.raise_for_status()
            return response.json()
        except requests.exceptions.RequestException as e:
            print(f"[ERROR] Attempt {attempt + 1}: {e}")
            time.sleep(wait_s * (2 ** attempt))  # Exponential backoff
    return None  # Return None if all attempts fail

# Function to sanitize filenames
def sanitize_filename(name):
    return re.sub(r'[^a-zA-Z0-9_.-]', '_', name)

def search_unpaywall_by_title(keywords, email, is_oa=None, max_pages=5, output_xlsx="unpaywall_results.xlsx", wait_s=1.0):
    wb = Workbook()
    sheet_all = wb.active
    sheet_all.title = "All"
    sheet_all.append(["doi", "title", "publisher", "year", "oa_status", "best_oa_location_url", "journal_name"])

    for page_num in range(1, max_pages + 1):
        params = {"query": keywords, "email": email, "page": page_num}
        if is_oa is not None:
            params["is_oa"] = str(is_oa).lower()

        print(f"[INFO] Fetching page {page_num}...")
        data = retry_request(API_BASE_URL, params)
        if not data or "results" not in data:
            print(f"[INFO] No more articles found. Stopping at page {page_num}.")
            break

        for result in data["results"]:
            doi_obj = result.get("response", {})
            best_oa_loc = doi_obj.get("best_oa_location") or {}  # Ensure it's a dictionary
            best_oa_url = best_oa_loc.get("url_for_pdf") or best_oa_loc.get("url")

            sheet_all.append([
                doi_obj.get("doi"),
                doi_obj.get("title"),
                doi_obj.get("publisher"),
                doi_obj.get("year"),
                doi_obj.get("oa_status"),
                best_oa_url,
                doi_obj.get("journal_name")
            ])
        time.sleep(wait_s)

    wb.save(output_xlsx)
    print(f"[DONE] Results saved to '{output_xlsx}'.")

def filter_oa_status_not_closed(excel_file):
    wb = openpyxl.load_workbook(excel_file)
    sheet_all = wb["All"]

    if "Accessible" in wb.sheetnames:
        del wb["Accessible"]
    sheet_accessible = wb.create_sheet("Accessible")
    sheet_accessible.append([cell.value for cell in sheet_all[1]])

    oa_status_index = 4
    for row in sheet_all.iter_rows(min_row=2, values_only=True):
        if row[oa_status_index] != "closed":
            sheet_accessible.append(row)

    wb.save(excel_file)
    print(f"[INFO] Filtered data saved in 'Accessible' sheet.")

def download_paper(row, directory):
    doi, url = row[0], row[5]
    if not url:
        return f"[INFO] No URL for DOI '{doi}', skipping."
    
    safe_doi = sanitize_filename(doi)
    local_pdf_path = os.path.join(directory, f"{safe_doi}.pdf")

    try:
        response = requests.get(url, stream=True, timeout=30)
        content_type = response.headers.get("Content-Type", "").lower()
        disposition = response.headers.get("Content-Disposition", "")
        is_pdf = content_type.startswith("application/pdf") or ".pdf" in disposition

        if response.status_code == 200 and is_pdf:
            with open(local_pdf_path, "wb") as f:
                for chunk in response.iter_content(chunk_size=8192):
                    f.write(chunk)
            return f"[INFO] Downloaded: {local_pdf_path}"
        else:
            return f"[WARNING] Not a PDF or download failed. URL={url}"
    except requests.exceptions.RequestException as e:
        return f"[ERROR] Download failed for DOI '{doi}': {e}"

def download_papers_from_accessible(excel_file, directory="Papers"):
    wb = openpyxl.load_workbook(excel_file)
    if "Accessible" not in wb.sheetnames:
        print("[WARNING] 'Accessible' sheet not found. No downloads attempted.")
        return

    os.makedirs(directory, exist_ok=True)
    sheet_accessible = wb["Accessible"]
    rows = [row for row in sheet_accessible.iter_rows(min_row=2, values_only=True)]

    with ThreadPoolExecutor(max_workers=5) as executor:
        results = executor.map(lambda row: download_paper(row, directory), rows)

    for result in results:
        print(result)

if __name__ == "__main__":
    query_keywords = "rembrandt paintings"
    search_unpaywall_by_title(query_keywords, EMAIL, is_oa=None, max_pages=3, output_xlsx="vg_paintings.xlsx", wait_s=1.0)
    filter_oa_status_not_closed("rembrandt_paintings.xlsx")
    download_papers_from_accessible("rembrandt_paintings.xlsx", directory="Rembrandt")
