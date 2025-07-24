import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Define the desired column order
desired_columns = [
    "display_name",
    "id",
    "subfield",
    "field",
    "domain",
    "works_count",
    "cited_by_count",
    "description",
    "keywords",
]

# Read the Excel file and reorder columns
df = pd.read_excel("openalex_topics.xlsx", engine="openpyxl")
df = df[desired_columns]

# Save the reordered data to a new file
output_filename = "openalex_topics_formatted.xlsx"
df.to_excel(output_filename, index=False, engine="openpyxl")

# Open the new workbook to adjust column widths
wb = load_workbook(output_filename)
ws = wb.active

# Auto-adjust column widths (except for description and keywords)
for col in ws.columns:
    header = col[0].value
    if header is None:
        continue
    # Skip the description and keywords columns
    if header.lower() in ["description", "keywords"]:
        continue
    max_length = len(str(header))
    for cell in col:
        if cell.value is not None:
            max_length = max(max_length, len(str(cell.value)))
    adjusted_width = max_length + 2  # add some padding
    col_letter = get_column_letter(col[0].column)
    ws.column_dimensions[col_letter].width = adjusted_width

wb.save(output_filename)
print(f"Formatted Excel file saved as {output_filename}")
