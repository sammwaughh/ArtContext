import os
import nltk
import numpy as np
import pandas as pd
import torch
import time
import re
from openpyxl import load_workbook
from sentence_transformers import SentenceTransformer, util

def clean_text(text):
    # Remove illegal control characters that Excel cannot handle.
    return re.sub(r'[\x00-\x08\x0B-\x0C\x0E-\x1F]', '', text)

def find_matching_paintings(paintings_range, painter_name):
    # Try an exact full match (case-insensitive).
    mask_exact = paintings_range["Creator"].str.lower() == painter_name.lower()
    if mask_exact.sum() > 0:
        print(f"Found {mask_exact.sum()} paintings with exact full match for '{painter_name}'.")
        return paintings_range[mask_exact]
    
    # Try matching by first name exactly.
    first_name = painter_name.split()[0]
    mask_first = paintings_range["Creator"].str.lower() == first_name.lower()
    if mask_first.sum() > 0:
        print(f"No full match found. Found {mask_first.sum()} paintings with first name '{first_name}'.")
        return paintings_range[mask_first]
    
    # Try matching by last name exactly.
    last_name = painter_name.split()[-1]
    mask_last = paintings_range["Creator"].str.lower() == last_name.lower()
    if mask_last.sum() > 0:
        print(f"No full or first name match found. Found {mask_last.sum()} paintings with last name '{last_name}'.")
        return paintings_range[mask_last]
    
    print(f"No paintings found for variations of '{painter_name}'.")
    return pd.DataFrame()

def extract_sentences_from_markdown(file_path):
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    # Remove code blocks, inline code, markdown links, images, and remaining markdown symbols.
    content = re.sub(r'```[\s\S]*?```', '', content)
    content = re.sub(r'`[^`]*`', '', content)
    content = re.sub(r'\[([^\]]+)\]\([^)]+\)', r'\1', content)
    content = re.sub(r'!\[[^\]]*\]\([^)]+\)', '', content)
    content = re.sub(r'[#>*_~\-]', '', content)
    sentences = nltk.sent_tokenize(content)
    return [s.strip() for s in sentences if len(s.split()) > 3]

def build_context_sentences(sentences):
    """
    For each sentence, create a 3-sentence context string by concatenating:
    the previous sentence, the current sentence, and the next sentence (if available),
    using " ||| " as a delimiter.
    """
    candidate_contexts = []
    for i, sentence in enumerate(sentences):
        context = []
        if i > 0:
            context.append(sentences[i-1])
        context.append(sentence)
        if i < len(sentences) - 1:
            context.append(sentences[i+1])
        candidate_contexts.append(" ||| ".join(context))
    return candidate_contexts

def main():
    nltk.download('punkt')
    nltk.download('punkt_tab')
    
    start_time = time.perf_counter()
    
    # Load paintings.xlsx and painters.xlsx.
    paintings_df = pd.read_excel("paintings.xlsx")
    painters_df = pd.read_excel("painters.xlsx")
    cache_dir = "cache"
    
    # Specify an inclusive range of paintings (rows in paintings.xlsx) to process (0-indexed).
    painting_start_index = 0
    painting_end_index = 40000  # e.g., process rows 0 through 40000 inclusive.
    
    # Specify an inclusive range of painters (rows in painters.xlsx) to process (0-indexed).
    painter_start_index = 0
    painter_end_index = 454  # Adjust as needed.
    
    paintings_range = paintings_df.iloc[painting_start_index:painting_end_index+1]
    painters_subset = painters_df.iloc[painter_start_index:painter_end_index+1]
    
    # Load the existing workbook with openpyxl to preserve formatting.
    wb = load_workbook("paintings.xlsx")
    ws = wb.active

    # Build a dictionary mapping header names to their column numbers (from header row 1).
    header_map = {}
    for cell in ws[1]:
        if cell.value is not None:
            header_map[cell.value] = cell.column

    device = "mps" if torch.backends.mps.is_available() else "cpu"
    print(f"Using device: {device}")
    print("Loading Sentence-BERT model...")
    model = SentenceTransformer('paraphrase-MiniLM-L6-v2', device=device)
    
    total_processed = 0

    # Iterate over each painter in the specified painters range.
    for _, painter_row in painters_subset.iterrows():
        painter_name = painter_row["Artist"]
        artist_dir_name = painter_row["Query String"]
        print(f"\nProcessing painter: {painter_name} with directory mapping: {artist_dir_name}")
        
        cache_file = os.path.join(cache_dir, f"{artist_dir_name}_cache.pt")
        if not os.path.exists(cache_file):
            print(f"No cache file for painter {painter_name} at {cache_file}. Skipping.")
            continue
        print(f"Loading cache from {cache_file}...")
        data = torch.load(cache_file)
        candidate_contexts_all = data["candidate_contexts"]
        candidate_sentences_all = data["candidate_sentences"]
        candidate_embeddings = data["candidate_embeddings"].to(device)
        
        matching_paintings = find_matching_paintings(paintings_range, painter_name)
        if matching_paintings.empty:
            continue
        
        # Process each matching painting (no per-painting prints)
        for idx, row in matching_paintings.iterrows():
            try:
                painting_title = row["Title"]
                artist = row["Creator"]
            except IndexError:
                continue

            query = f"Describe the visual content of '{painting_title}' by {artist}."
            query_embedding = model.encode(query, convert_to_tensor=True)
            
            chunk_size = 10000
            cosine_scores_list = []
            num_candidates = candidate_embeddings.shape[0]
            for i in range(0, num_candidates, chunk_size):
                chunk = candidate_embeddings[i:i+chunk_size]
                scores_chunk = util.cos_sim(query_embedding, chunk)[0]
                cosine_scores_list.append(scores_chunk.cpu().numpy())
            cosine_scores_np = np.concatenate(cosine_scores_list)
            best_index = np.argmax(cosine_scores_np)
            best_sequence = candidate_contexts_all[best_index]
            parts = best_sequence.split(" ||| ")
            if len(parts) != 3:
                parts = [best_sequence, "", ""]
            parts = [clean_text(part) for part in parts]
            
            # Update the corresponding row in the workbook.
            excel_row = idx + 2
            ws.cell(row=excel_row, column=header_map["Sentence 1"], value=parts[0])
            ws.cell(row=excel_row, column=header_map["Sentence 2"], value=parts[1])
            ws.cell(row=excel_row, column=header_map["Sentence 3"], value=parts[2])
            
            total_processed += 1

    wb.save("paintings.xlsx")
    elapsed_time = time.perf_counter() - start_time
    elapsed_rounded = round(elapsed_time, 1)
    rate = round(elapsed_time / total_processed, 1) if total_processed > 0 else float('inf')
    
    print("\nUpdated paintings.xlsx with the best sequences for all processed paintings.")
    print(f"Processed {total_processed} paintings in {elapsed_rounded} seconds. Rate: {rate} seconds per painting.")

if __name__ == "__main__":
    main()
