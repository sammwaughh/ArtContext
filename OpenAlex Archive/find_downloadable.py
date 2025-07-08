import pandas as pd
import json
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def safe_json_parse(cell_value):
    """
    If the cell value is already a dict or list, return it.
    Otherwise, try to parse it as JSON.
    """
    if isinstance(cell_value, (dict, list)):
        return cell_value
    try:
        return json.loads(cell_value)
    except Exception:
        return None

def get_candidate_links(row):
    """
    Given a row (from the Filtered sheet), extract candidate download links
    from several fields.
    """
    candidates = []
    
    # Best OA Location: prefer its pdf_url then landing_page_url
    best_oa = safe_json_parse(row.get('best_oa_location'))
    if best_oa and isinstance(best_oa, dict):
        if best_oa.get('pdf_url'):
            candidates.append(best_oa['pdf_url'])
        elif best_oa.get('landing_page_url'):
            candidates.append(best_oa['landing_page_url'])
    
    # Open Access: check for oa_url
    oa = safe_json_parse(row.get('open_access'))
    if oa and isinstance(oa, dict):
        if oa.get('oa_url'):
            candidates.append(oa['oa_url'])
    
    # Primary Location: pdf_url then landing_page_url
    primary = safe_json_parse(row.get('primary_location'))
    if primary and isinstance(primary, dict):
        if primary.get('pdf_url'):
            candidates.append(primary['pdf_url'])
        elif primary.get('landing_page_url'):
            candidates.append(primary['landing_page_url'])
    
    # Locations: iterate over each and check for pdf_url then landing_page_url
    locs = safe_json_parse(row.get('locations'))
    if locs and isinstance(locs, list):
        for loc in locs:
            if isinstance(loc, dict):
                if loc.get('pdf_url'):
                    candidates.append(loc['pdf_url'])
                elif loc.get('landing_page_url'):
                    candidates.append(loc['landing_page_url'])
    
    # Remove duplicates while preserving order
    return list(dict.fromkeys(candidates))

def get_best_and_backup_links(row):
    """
    Returns a tuple (best_link, backup_link) for a given row.
    
    Strategy:
      - Among candidate links, choose as best_link the first link that contains '.pdf'
        (case-insensitive). If none, use the first candidate.
      - Use the next candidate (if any) as backup_link.
    """
    candidates = get_candidate_links(row)
    best_link = ""
    backup_link = ""
    
    # Prefer a candidate containing ".pdf" (case-insensitive)
    for link in candidates:
        if '.pdf' in link.lower():
            best_link = link
            break
    # If no .pdf link found, use first candidate if available.
    if not best_link and candidates:
        best_link = candidates[0]
    
    # For backup, take the next candidate that is different from best_link.
    for link in candidates:
        if link != best_link:
            backup_link = link
            break
            
    return best_link, backup_link

def create_downloadable_sheet(excel_file):
    """
    Reads the "Filtered" sheet from the given Excel file and creates a new sheet called 
    "Downloadable" containing the columns:
      - Title
      - Relevance Score
      - Best Link
      - Back Up Link
      - OpenAlexID
      
    The Best Link is chosen using get_best_and_backup_links() and the Back Up Link is the second candidate.
    All columns in the Downloadable sheet are set to a fixed width of 35.
    """
    # Read the "Filtered" sheet into a DataFrame
    df = pd.read_excel(excel_file, sheet_name="Filtered", engine="openpyxl")
    
    downloadable_data = []
    for idx, row in df.iterrows():
        title = row.get('title', '')
        relevance = row.get('relevance_score', '')
        openalex_id = row.get('id', '')
        best_link, backup_link = get_best_and_backup_links(row)
        downloadable_data.append({
            'Title': title,
            'Relevance Score': relevance,
            'Best Link': best_link,
            'Back Up Link': backup_link,
            'OpenAlexID': openalex_id
        })
    
    downloadable_df = pd.DataFrame(downloadable_data)
    
    # Open the workbook with openpyxl
    wb = load_workbook(excel_file)
    # Remove existing "Downloadable" sheet if present.
    if "Downloadable" in wb.sheetnames:
        ws_del = wb["Downloadable"]
        wb.remove(ws_del)
    ws_download = wb.create_sheet(title="Downloadable")
    
    # Write header row.
    headers = list(downloadable_df.columns)
    ws_download.append(headers)
    
    # Write data rows.
    for row in downloadable_df.itertuples(index=False, name=None):
        ws_download.append(list(row))
    
    # Set all column widths in the Downloadable sheet to 35.
    for col in ws_download.columns:
        col_letter = get_column_letter(col[0].column)
        ws_download.column_dimensions[col_letter].width = 35
    
    wb.save(excel_file)
    print(f"Downloadable sheet created and saved in {excel_file}")

if __name__ == "__main__":
    excel_file = "vermeer_works.xlsx"
    create_downloadable_sheet(excel_file)
