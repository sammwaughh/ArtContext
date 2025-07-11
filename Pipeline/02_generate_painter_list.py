#!/usr/bin/env python3
"""
Build painters.xlsx from paintings.xlsx.

Input  (script 1 output)
------------------------
paintings.xlsx – sheet “Paintings”, column “Creator”

Output (for script 3)
---------------------
painters.xlsx – sheet “Painters” with two columns
• Artist        – original creator name
• Query String  – lower-case ASCII, stripped of brackets / accents

The sheet is auto-sized so script 3 can read it unchanged.
"""

from __future__ import annotations

import unicodedata
from pathlib import Path
from typing import NoReturn

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ──────────────────────────── constants ───────────────────────────────────────

SOURCE_XLSX: Path = Path("paintings.xlsx")
DEST_XLSX: Path = Path("painters.xlsx")


def to_query(s: str) -> str:
    """
    Return a search-friendly version of *s*:
    lower-case, accents removed, text inside “()” dropped.
    """
    s = s.split("(")[0].strip()  # drop bracketed nick-names etc.
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()
    return " ".join(s.lower().split())


def autosize(ws: Worksheet) -> None:
    """Resize every column width to fit its longest cell."""
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = max(len(str(cell.value)) for cell in col if cell.value is not None)
        ws.column_dimensions[letter].width = max_len + 2


def main() -> NoReturn:
    """Extract creators from paintings.xlsx and save painters.xlsx."""

    source_path: Path = SOURCE_XLSX
    if not source_path.is_file():
        raise FileNotFoundError(f"{SOURCE_XLSX} (output of script 1) not found")

    # read the “Creator” column only
    df_paintings = pd.read_excel(
        source_path,
        sheet_name="Paintings",
        usecols=["Creator"],
    )
    creators = (
        df_paintings["Creator"].dropna().astype(str).str.strip().unique().tolist()
    )

    # preserve original order of appearance
    artists = pd.Series(creators).drop_duplicates().tolist()

    data = [{"Artist": a, "Query String": to_query(a)} for a in artists]
    df_out = pd.DataFrame(data)

    with pd.ExcelWriter(DEST_XLSX, engine="openpyxl") as writer:
        df_out.to_excel(writer, index=False, sheet_name="Painters")

    # adjust widths
    wb = load_workbook(DEST_XLSX)
    ws = wb["Painters"]
    autosize(ws)
    wb.save(DEST_XLSX)
    print(f"Created {DEST_XLSX} with {len(df_out)} unique artists.")


if __name__ == "__main__":
    main()
