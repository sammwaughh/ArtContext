import os
import re
import string
import pandas as pd
import pypdfium2
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

PDFS_DIR = "PDFs"
EXCEL_DIR = "ExcelFiles"

def normalize_title(title):
    """Convert a title to lowercase and remove punctuation for matching."""
    translator = str.maketrans('', '', string.punctuation)
    return title.strip().lower().translate(translator)

def extract_pdf_metadata(pdf_path):
    """Extract metadata from a PDF file's filename."""
    filename = os.path.basename(pdf_path)
    base, _ = os.path.splitext(filename)
    match = re.match(r"(\d+)-(.+)", base)
    work_name = match.group(2).strip() if match else base.strip()
    file_size = os.path.getsize(pdf_path)
    try:
        _ = pypdfium2.PdfDocument(pdf_path)
        file_type = "PDF"
    except Exception as e:
        file_type = f"Error: {e}"
    return {
        "full name": filename,
        "work_name": work_name,  # temporary for lookup
        "file size": file_size,  # in bytes
        "type": file_type
    }

# Specify the inclusive range (1-indexed) of rows to process from painters.xlsx.
start_row = 301
end_row = 451

# Read the painters.xlsx file and get the list of painter names from the "Query String" column.
df_painters = pd.read_excel("painters.xlsx")
painter_names = df_painters.loc[start_row - 1:end_row - 1, "Query String"].tolist()

# Process only the painters specified in the Excel file.
for painter in painter_names:
    # Use the painter name from the Excel file to determine the input folder and Excel file.
    painter_folder = os.path.join(PDFS_DIR, painter)
    if not os.path.isdir(painter_folder):
        print(f"Folder not found for painter '{painter}': {painter_folder}")
        continue

    pdf_metadata = []
    for file in os.listdir(painter_folder):
        if file.lower().endswith('.pdf'):
            pdf_path = os.path.join(painter_folder, file)
            pdf_metadata.append(extract_pdf_metadata(pdf_path))
    if not pdf_metadata:
        continue

    # Excel file is expected to be named "{painter}_works.xlsx" in the ExcelFiles directory.
    excel_file = os.path.join(EXCEL_DIR, f"{painter}_works.xlsx")
    if not os.path.exists(excel_file):
        print(f"Excel file not found for painter '{painter}': {excel_file}")
        continue

    try:
        df_downloadable = pd.read_excel(excel_file, sheet_name="Downloadable")
    except Exception as e:
        print(f"Error reading 'Downloadable' sheet in {excel_file}: {e}")
        continue

    # Normalize the Title column for matching.
    df_downloadable["norm_title"] = df_downloadable["Title"].astype(str).apply(normalize_title)

    # For each PDF, look up its work name (normalized) to get its Relevance Score.
    for item in pdf_metadata:
        norm_work = normalize_title(item["work_name"])
        match = df_downloadable.loc[df_downloadable["norm_title"] == norm_work, "Relevance Score"]
        item["relevance score"] = match.iloc[0] if not match.empty else None

    # Convert file size from bytes to kilobytes.
    df_markdown = pd.DataFrame(pdf_metadata)
    df_markdown["file size"] = df_markdown["file size"].apply(lambda x: round(x / 1024, 2))
    
    # Keep only rows where type is exactly "PDF"
    df_markdown = df_markdown[df_markdown["type"] == "PDF"]

    # Sort the DataFrame by relevance score (descending order).
    df_markdown.sort_values(by="relevance score", ascending=False, inplace=True)

    # Remove duplicate entries based on the normalized full name.
    # Here we remove any numeric prefix before the first dash.
    df_markdown["norm_full_name"] = df_markdown["full name"].apply(
        lambda x: x.split("-", 1)[-1].strip().lower() if "-" in x else x.strip().lower()
    )
    df_markdown = df_markdown.drop_duplicates(subset=["norm_full_name"], keep="first")
    df_markdown = df_markdown.drop(columns=["norm_full_name"])

    # Keep only the required columns.
    df_markdown = df_markdown[["full name", "relevance score", "file size", "type"]]

    # Write/update the "For Markdown" sheet in the painter's works Excel file.
    try:
        with pd.ExcelWriter(excel_file, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            df_markdown.to_excel(writer, sheet_name="For Markdown", index=False)
        print(f"Updated {excel_file} with sheet 'For Markdown'.")
    except Exception as e:
        print(f"Error writing to {excel_file}: {e}")
        continue

    # Adjust column widths for readability.
    try:
        wb = load_workbook(excel_file)
        ws = wb["For Markdown"]
        for col_cells in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2
        wb.save(excel_file)
    except Exception as e:
        print(f"Error adjusting column widths in {excel_file}: {e}")
