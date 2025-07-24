import json
import time

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Set your email here
my_email = "xlct43@durham.ac.uk"


def convert_to_str(value):
    """Convert dictionaries or lists to a JSON string; otherwise, return the value."""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def fetch_works(painter_name):
    """
    Fetch all works from OpenAlex that meet the following criteria:
      - language is English (language:en)
      - the work is open access (is_oa:true)
      - the work has one of the specified topic IDs in its topics list
      - the work's searchable metadata mentions the painter_name
    """
    base_url = "https://api.openalex.org/works"
    # List of topic IDs separated by a vertical bar (OR logic)
    topics_ids = (
        "T14092|T14191|T12372|T14469|T12680|T14366|T13922|T12444|"
        "T13133|T12179|T13342|T12632|T14002|T14322"
    )
    # Build the filter: English language, open access, and topics.id equal to one of these IDs.
    filter_str = f"language:en,is_oa:true,topics.id:{topics_ids}"

    # Use cursor-based pagination; start with cursor set to "*"
    cursor = "*"
    per_page = 200  # maximum allowed per page is 200
    works = []
    page_count = 0

    # Set custom headers (including a User-Agent)
    headers = {"User-Agent": f"MyPythonClient (mailto:{my_email})"}

    print(f"Starting query for works mentioning '{painter_name}'...")

    while cursor:
        page_count += 1
        print(f"Querying page {page_count}...")
        params = {
            "filter": filter_str,
            "search": painter_name,
            "per_page": per_page,
            "cursor": cursor,
            "mailto": my_email,
        }

        while True:
            response = requests.get(base_url, params=params, headers=headers)
            if response.status_code == 429:
                print("Rate limit exceeded. Sleeping for 60 seconds...")
                time.sleep(60)
            else:
                break

        if response.status_code != 200:
            print(f"Error: Received status code {response.status_code}")
            print("Response content:")
            print(response.text)
            break

        data = response.json()
        results = data.get("results", [])
        works.extend(results)
        print(f"Page {page_count}: Retrieved {len(results)} works.")

        cursor = data.get("meta", {}).get("next_cursor")
        if not cursor:
            print("No further pages found.")
            break

        time.sleep(1)

    print(
        f"Query complete. Total pages: {page_count}. Total works retrieved: {len(works)}"
    )
    return works


def save_to_excel(data, filename):
    """
    Save the list of work dictionaries to an Excel file using pandas,
    then create a second sheet with only a subset of columns.
    The main sheet is auto-adjusted for column widths,
    and the "Filtered" sheet has all column widths set to 35.
    """
    # Convert the list of dictionaries to a DataFrame.
    df = pd.DataFrame(data)

    # Write the main sheet (with all columns)
    df.to_excel(filename, index=False, engine="openpyxl")

    # Open the workbook to adjust column widths on the main sheet.
    wb = load_workbook(filename)
    ws = wb.active  # Main sheet (default "Sheet1")
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_length = 0
        for cell in col:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2
    print("Main sheet saved and formatted.")

    # Define the subset of columns for the second sheet.
    selected_columns = [
        "title",
        "relevance_score",
        "id",
        "doi",
        "primary_location",
        "type",
        "open_access",
        "locations",
        "best_oa_location",
    ]
    # Create a new DataFrame with only the selected columns.
    df_subset = df[selected_columns].copy()

    # Create a new sheet named "Filtered" and write the subset data.
    ws_subset = wb.create_sheet(title="Filtered")
    # Write header row.
    ws_subset.append(selected_columns)
    # Write data rows.
    for row in df_subset.itertuples(index=False, name=None):
        new_row = [convert_to_str(value) for value in row]
        ws_subset.append(new_row)

    # Set all column widths in the "Filtered" sheet to 35.
    for col in ws_subset.columns:
        col_letter = get_column_letter(col[0].column)
        ws_subset.column_dimensions[col_letter].width = 35

    wb.save(filename)
    print(f"Excel file with both sheets saved and formatted: {filename}")


def main():
    painter_name = "vermeer"  # Change this variable to search for a different painter.
    works = fetch_works(painter_name)
    filename = f"{painter_name.lower()}_works.xlsx"
    save_to_excel(works, filename)


if __name__ == "__main__":
    main()
