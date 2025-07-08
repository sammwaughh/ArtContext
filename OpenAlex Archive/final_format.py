from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

input_filename = "openalex_topics_cleaned.xlsx"
output_filename = "openalex_topics_cleaned_formatted.xlsx"

# Load workbook and select active worksheet
wb = load_workbook(input_filename)
ws = wb.active

# Loop through each column (assumes header is in the first row)
for col in ws.columns:
    header = col[0].value
    if header and header.strip().lower() in ["description", "keywords"]:
        continue  # Skip adjusting these columns

    max_length = len(str(header)) if header else 0
    for cell in col:
        if cell.value is not None:
            max_length = max(max_length, len(str(cell.value)))
    adjusted_width = max_length + 2  # add padding
    col_letter = get_column_letter(col[0].column)
    ws.column_dimensions[col_letter].width = adjusted_width

# Save the formatted workbook
wb.save(output_filename)
print(f"Formatted file saved as {output_filename}")
