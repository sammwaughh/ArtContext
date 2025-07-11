#!/usr/bin/env python3
"""
Stage 3 of the pipeline – harvest OpenAlex literature for a set of painters,
store the metadata in Excel and download the relevant PDFs.

CLI
---
python 03_openalex_download.py                 # all painters
python 03_openalex_download.py -r 150          # first 150
python 03_openalex_download.py -r 350:420      # rows 350 – 420 (1-based, inclusive)

Style
-----
• PEP 526 type annotations
• SCREAMING_SNAKE_CASE constants
• Modern Python 3.9+ syntax
"""

from __future__ import annotations

import argparse
import asyncio
import json
import logging
import os
import re
import threading
import time
from logging.handlers import TimedRotatingFileHandler
from pathlib import Path
from typing import Dict, List, Tuple

import httpx
import pandas as pd
import psutil
import requests
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from requests.adapters import HTTPAdapter
from tqdm import tqdm
from urllib3.util.retry import Retry

# ──────────────────────────── configuration ───────────────────────────────────

CONTACT_EMAIL: str = "xlct43@durham.ac.uk"

# ── Excel folders ------------------------------------------------------------
EXCEL_DIR: Path = Path("Excel-Files")  # root for every workbook
WORKS_DIR: Path = EXCEL_DIR / "Painter-Works-Metadata"  # per-painter workbooks
PAINTERS_XLSX: Path = EXCEL_DIR / "painters.xlsx"  # input list

# ── Output PDFs --------------------------------------------------------------
# keep them inside the Pipeline package: “…/Pipeline/PDFs/”
PDF_DIR: Path = Path(__file__).resolve().parent / "PDFs"

# ── Logs ---------------------------------------------------------------------
LOG_ROOT: Path = Path("logs")
PAINTER_LOG_DIR: Path = LOG_ROOT / "painter-logs"
LOG_DIRS: Tuple[Path, ...] = (
    LOG_ROOT / "fetch-logs",
    LOG_ROOT / "excel-logs",
    LOG_ROOT / "download-logs",
    LOG_ROOT / "cpu-logs",
)

TOPIC_IDS: str = (
    "T14092|T14191|T12372|T14469|T12680|T14366|T13922|T12444|"
    "T13133|T12179|T13342|T12632|T14002|T14322"
)

DEFAULT_PER_PAGE: int = 200  # OpenAlex page size
PDF_CHUNK_SIZE: int = 8_192  # bytes
CPU_SAMPLE_SEC: int = 1
MAX_RETRIES: int = 3
BACKOFF_FACTOR: float = 0.3
SESSION_TIMEOUT_SEC: int = 20
MAX_WORKERS_FACTOR: int = 2  # ≈ workers = CPUs × factor
MIN_WORKERS: int = 4

START_ROW: int = 1  # default CLI range start (1-based)
END_ROW: int | None = None  # default CLI range end   (inclusive)
DEFAULT_SLEEP_SEC: float = 1.0  # politeness delay between API calls

# ──────────────────────────── helpers & utils ─────────────────────────────────


def _make_session() -> requests.Session:
    """Return a shared, retry-enabled HTTP session."""
    retry: Retry = Retry(
        total=MAX_RETRIES,
        backoff_factor=BACKOFF_FACTOR,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["HEAD", "GET", "OPTIONS"],
        raise_on_status=False,
    )
    sess: requests.Session = requests.Session()
    adapter: HTTPAdapter = HTTPAdapter(max_retries=retry)
    sess.mount("https://", adapter)
    sess.mount("http://", adapter)
    sess.headers.update({"User-Agent": f"ArtContext/0.1 (mailto:{CONTACT_EMAIL})"})
    return sess


SESSION: requests.Session = _make_session()


def _setup_logger() -> logging.Logger:
    """One rotating root logger (midnight UTC)."""
    log_dir: Path = Path("logs")
    log_dir.mkdir(exist_ok=True)
    handler = TimedRotatingFileHandler(
        log_dir / "openalex.log", when="midnight", utc=True, backupCount=7
    )
    handler.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
    root = logging.getLogger("openalex")
    root.setLevel(logging.INFO)
    root.addHandler(handler)
    return root


LOGGER: logging.Logger = _setup_logger()


def _safe_json_parse(value: str | Dict | List) -> Dict | List | None:
    """`json.loads` that survives invalid input."""
    if isinstance(value, (dict, list)):
        return value
    try:
        return json.loads(value)
    except Exception:  # noqa: BLE001
        return None


def _sanitize_filename(name: str) -> str:
    """Remove characters that are invalid on most OS filesystems."""
    return re.sub(r'[\\/*?:"<>|]', "", name).strip()


# ──────────────────────────── OpenAlex access ─────────────────────────────────


def fetch_works(
    painter: str,
    per_page: int,
    sleep_sec: float,
) -> List[Dict]:
    """
    Retrieve every OpenAlex work that is
    English, open-access, within TOPIC_IDS and mentions *painter*.
    """
    base_url: str = "https://api.openalex.org/works"
    query_filter: str = f"language:en,is_oa:true,topics.id:{TOPIC_IDS}"
    cursor: str | None = "*"
    works: List[Dict] = []
    page: int = 0

    LOGGER.info("Query start – painter=%s", painter)
    pbar = tqdm(desc=f"{painter} pages", unit="page")

    while cursor:
        page += 1
        pbar.update(1)
        params: Dict[str, str | int | float] = {
            "filter": query_filter,
            "search": painter,
            "per_page": per_page,
            "cursor": cursor,
            "mailto": CONTACT_EMAIL,
        }
        while True:
            resp = SESSION.get(base_url, params=params, timeout=SESSION_TIMEOUT_SEC)
            if resp.status_code != 429:
                break
            retry_after = int(resp.headers.get("Retry-After", "60"))
            LOGGER.warning("429 – sleeping %d s", retry_after)
            time.sleep(retry_after)

        if resp.status_code != 200:  # give up – log & abort
            LOGGER.error("HTTP %s – %s", resp.status_code, resp.text[:200])
            break

        data: Dict = resp.json()
        page_results: List[Dict] = data.get("results", [])
        works.extend(page_results)
        cursor = data.get("meta", {}).get("next_cursor")

        LOGGER.info("%s page %d – %d works", painter, page, len(page_results))
        time.sleep(sleep_sec)

    pbar.close()
    LOGGER.info("%s finished – pages=%d  works=%d", painter, page, len(works))
    # Early relevance filter
    return [w for w in works if w.get("relevance_score", 0) > 1]


# ───────────────────────── Excel creation ─────────────────────────────────────

MAIN_COLS: Tuple[str, ...] = (
    "title",
    "relevance_score",
    "id",
    "doi",
    "primary_location",
    "type",
    "open_access",
    "locations",
    "best_oa_location",
)


def _get_candidate_links(row: pd.Series) -> List[str]:
    """Collect all distinct OA/PDF URLs from the various location blobs."""
    candidates: List[str] = []
    best_oa = _safe_json_parse(row["best_oa_location"])
    if isinstance(best_oa, dict):
        candidates += [best_oa.get("pdf_url"), best_oa.get("landing_page_url")]

    oa = _safe_json_parse(row["open_access"])
    if isinstance(oa, dict):
        candidates.append(oa.get("oa_url"))

    primary = _safe_json_parse(row["primary_location"])
    if isinstance(primary, dict):
        candidates += [primary.get("pdf_url"), primary.get("landing_page_url")]

    for loc in _safe_json_parse(row["locations"]) or []:
        if isinstance(loc, dict):
            candidates += [loc.get("pdf_url"), loc.get("landing_page_url")]

    return [c for c in dict.fromkeys(candidates) if c]  # dedupe & drop None


def _best_and_backup(row: pd.Series) -> Tuple[str, str]:
    """Return “best” PDF link plus a backup (may be empty strings)."""
    for_best: List[str] = _get_candidate_links(row)
    best: str = ""
    backup: str = ""
    for link in for_best:
        if ".pdf" in link.lower():
            best = link
            break
    if not best and for_best:
        best = for_best[0]
    for link in for_best:
        if link != best:
            backup = link
            break
    return best or "", backup or ""


def create_excel(painter: str, works: List[Dict]) -> Path:
    """Write three sheets (Main, Filtered, Downloadable) and return the file path."""
    df_main: pd.DataFrame = (
        pd.DataFrame(works) if works else pd.DataFrame(columns=MAIN_COLS)
    )
    df_filtered: pd.DataFrame = df_main[MAIN_COLS].copy()

    rows_downloadable: List[Dict[str, str | float]] = []
    for _, row in df_filtered.iterrows():
        best, backup = _best_and_backup(row)
        rows_downloadable.append(
            {
                "Title": row.get("title", ""),
                "Relevance Score": row.get("relevance_score", 0.0),
                "Best Link": best,
                "Backup Link": backup,
                "OpenAlex ID": row.get("id", ""),
                "Type": row.get("type", ""),
            }
        )
    df_download: pd.DataFrame = pd.DataFrame(rows_downloadable)

    dest: Path = WORKS_DIR / f"{painter.lower()}_works.xlsx"
    with pd.ExcelWriter(dest, engine="openpyxl") as writer:
        df_main.to_excel(writer, index=False, sheet_name="Main")
        df_filtered.to_excel(writer, index=False, sheet_name="Filtered")
        df_download.to_excel(writer, index=False, sheet_name="Downloadable")

    # autosize columns (rough)
    wb = load_workbook(dest)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for column in ws.columns:
            width: int = max(len(str(cell.value)) for cell in column if cell.value) + 2
            ws.column_dimensions[get_column_letter(column[0].column)].width = min(
                width, 50
            )
    wb.save(dest)
    LOGGER.info("%s workbook saved (%d rows)", painter, len(df_main))
    return dest


# ───────────────────────── PDF download (async) ────────────────────────────


async def _grab_pdf(
    idx: int,
    title: str,
    best: str,
    backup: str,
    sem: asyncio.Semaphore,
    out_dir: Path,
) -> Tuple[str, float] | None:
    """Async download with automatic Retry-After handling."""
    filename = f"{idx}-{_sanitize_filename(title)}.pdf"
    path = out_dir / filename
    start = time.perf_counter()

    async with sem:
        async with httpx.AsyncClient(timeout=SESSION_TIMEOUT_SEC) as client:
            for url in (best, backup):
                if not url:
                    continue
                try:
                    r = await client.get(url, follow_redirects=True)
                    if r.status_code == 429:
                        retry_after = int(r.headers.get("Retry-After", "60"))
                        await asyncio.sleep(retry_after)
                        continue
                    if r.status_code != 200:
                        LOGGER.warning(
                            "Row %d – HTTP %s on %s", idx, r.status_code, url
                        )
                        continue
                    path.write_bytes(r.content)
                    return filename, round(time.perf_counter() - start, 3)
                except Exception as exc:  # noqa: BLE001
                    LOGGER.error("Row %d – %s on %s", idx, exc, url)
    return None


def _monitor_cpu(stop_evt: threading.Event, clog: logging.Logger) -> None:
    """Log instantaneous and average CPU usage until *stop_evt* is set."""
    samples: List[float] = []
    while not stop_evt.wait(CPU_SAMPLE_SEC):
        usage: float = psutil.cpu_percent()
        samples.append(usage)
        clog.info("CPU %.1f%%", usage)
    if samples:
        clog.info("Average CPU %.2f%%", sum(samples) / len(samples))


async def download_pdfs(
    excel_file: Path,
    painter: str,
    max_workers: int,
    dlog: logging.Logger,
    clog: logging.Logger,
) -> List[Tuple[str, float]]:
    """
    Parallel PDF download (relevance_score > 1).
    Returns list of (pdf_name, seconds).
    """
    df: pd.DataFrame = pd.read_excel(
        excel_file, sheet_name="Downloadable", engine="openpyxl"
    )
    df = df.drop_duplicates(subset=["Title"], keep="first")

    out_dir: Path = PDF_DIR / painter.lower()
    out_dir.mkdir(parents=True, exist_ok=True)

    stop_evt = threading.Event()
    threading.Thread(target=_monitor_cpu, args=(stop_evt, clog), daemon=True).start()

    sem = asyncio.Semaphore(max_workers)
    tasks = []
    for idx, row in enumerate(df.itertuples(index=False), start=1):
        tasks.append(
            _grab_pdf(idx, row[0], row[2], row[3], sem, out_dir)
        )  # Title / links
    results: List[Tuple[str, float]] = []
    for coro in tqdm(
        asyncio.as_completed(tasks), total=len(tasks), desc=f"{painter} PDFs"
    ):
        res = await coro
        if res:
            results.append(res)

    stop_evt.set()

    failed = {row[0] for row in df.itertuples(index=False)} - {r[0] for r in results}
    if failed:
        (out_dir / "failed_downloads.txt").write_text("\n".join(sorted(failed)))
    return results


def append_download_times(excel_file: Path, times: List[Tuple[str, float]]) -> None:
    """Append a ‘Download Times’ sheet to *excel_file* (overwrites if exists)."""
    if not times:
        return
    df: pd.DataFrame = pd.DataFrame(times, columns=["PDF Name", "Seconds"])
    df = df.sort_values("Seconds", ascending=False).reset_index(drop=True)

    wb = load_workbook(excel_file)
    if "Download Times" in wb.sheetnames:
        wb.remove(wb["Download Times"])
    ws = wb.create_sheet("Download Times")
    ws.append(["PDF Name", "Seconds"])
    for name, sec in df.itertuples(index=False):
        ws.append([name, sec])
    for column in ws.columns:
        width = max(len(str(c.value)) for c in column if c.value) + 2
        ws.column_dimensions[get_column_letter(column[0].column)].width = width
    wb.save(excel_file)


# ───────────────────────── painter wrapper ────────────────────────────────────


def process_painter(
    painter: str,
    max_workers: int,
    per_page: int,
    sleep_sec: float,
) -> None:
    """One-stop processing for a single painter."""
    ts: str = time.strftime("%Y%m%d-%H%M%S")
    dlog = _setup_logger("download", PAINTER_LOG_DIR / f"{painter}-{ts}.log")
    clog = _setup_logger("cpu", PAINTER_LOG_DIR / f"{painter}-{ts}.log")

    # fetch & store
    start_time: float = time.perf_counter()
    works = fetch_works(painter, per_page=per_page, sleep_sec=sleep_sec)
    if not works:
        print("  no works found – skipped")
        return

    excel_file: Path = create_excel(painter, works)
    times = asyncio.run(download_pdfs(excel_file, painter, max_workers, dlog, clog))
    append_download_times(excel_file, times)

    elapsed: float = time.perf_counter() - start_time
    print(f"  done in {elapsed:.1f} s  ({len(times)} PDFs)")


# ─────────────────────────── main entry point ────────────────────────────────


def _parse_range(text: str) -> Tuple[int, int | None]:
    """
    Convert “N” or “A:B” into (start, end)  – both 1-based, *end* inclusive.
    Validation is minimal; ValueError propagates to argparse.
    """
    if ":" in text:
        a, b = (int(x) for x in text.split(":", 1))
        return a, b
    return 1, int(text)


def main() -> None:
    """CLI front-door."""
    parser = argparse.ArgumentParser(
        description="Download OpenAlex PDFs for painters "
        "(page size, sleep time and workers are taken from the constants "
        "defined at the top of the file)."
    )
    parser.add_argument(
        "-r",
        "--rows",
        metavar="N | A:B",
        help="Rows in painters.xlsx (1-based).  E.g. 100  or  350:420.",
    )
    parser.add_argument(
        "-w",
        "--workers",
        type=int,
        help="Number of concurrent workers (default: auto-detect)",
    )
    parser.add_argument(
        "-p",
        "--per-page",
        type=int,
        default=DEFAULT_PER_PAGE,
        help="Results per page for OpenAlex API (default: %(default)d)",
    )
    parser.add_argument(
        "-s",
        "--sleep",
        type=float,
        default=DEFAULT_SLEEP_SEC,
        help="Sleep time between requests (default: %(default).1f s)",
    )
    args = parser.parse_args()

    start_row, end_row = (
        (START_ROW, END_ROW) if not args.rows else _parse_range(args.rows)
    )

    # ensure directories
    EXCEL_DIR.mkdir(exist_ok=True)
    WORKS_DIR.mkdir(exist_ok=True)
    PDF_DIR.mkdir(exist_ok=True)
    for d in (*LOG_DIRS, PAINTER_LOG_DIR):
        d.mkdir(parents=True, exist_ok=True)

    df_painters: pd.DataFrame = pd.read_excel(PAINTERS_XLSX)
    subset: pd.DataFrame = df_painters.iloc[
        start_row - 1 : end_row
    ]  # end_row may be None

    max_workers: int = max(MIN_WORKERS, os.cpu_count() * MAX_WORKERS_FACTOR)
    if args.workers:
        max_workers = max(1, args.workers)

    for i, row in enumerate(subset.itertuples(index=False), start=start_row):
        painter: str = row[1] if "Query String" in df_painters.columns else row[0]
        print(f"Row {i}", end="")
        process_painter(
            painter,
            max_workers=max_workers,
            per_page=args.per_page,
            sleep_sec=args.sleep,
        )


if __name__ == "__main__":
    main()
