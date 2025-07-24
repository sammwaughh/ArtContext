from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


def filter_bold_topic_name(
    input_file="topics_art_reordered.xlsx", output_file="topics_art_filtered.xlsx"
):
    # Load the source workbook and select the active sheet (assumed to be the merged sheet)
    wb = load_workbook(input_file)
    ws = wb.active

    # Read the header row to find the "Topic Name" column index
    header = [cell.value for cell in ws[1]]
    try:
        topic_name_col_index = (
            header.index("Topic Name") + 1
        )  # openpyxl columns are 1-indexed
    except ValueError:
        print("Error: 'Topic Name' column not found.")
        return

    # Create a new workbook with a single sheet named "topic_art_filtered"
    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.title = "topic_art_filtered"

    # Copy the header row to the new sheet
    new_ws.append(header)

    # Iterate over the data rows; check if the Topic Name cell is in bold
    for row in ws.iter_rows(min_row=2):
        topic_cell = row[
            topic_name_col_index - 1
        ]  # adjust for 0-indexing in the row list
        if topic_cell.font and topic_cell.font.bold:
            # Extract the cell values from the entire row
            row_values = [cell.value for cell in row]
            new_ws.append(row_values)

    # Auto-adjust column widths in the new sheet for better readability
    for col in new_ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_length = max(
            (len(str(cell.value)) for cell in col if cell.value is not None), default=0
        )
        new_ws.column_dimensions[col_letter].width = max_length + 2  # add some padding

    new_wb.save(output_file)
    print(
        f"Filtered rows saved in '{output_file}' under the sheet 'topic_art_filtered'."
    )


if __name__ == "__main__":
    filter_bold_topic_name()
