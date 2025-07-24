import datetime
import logging
import os
import re
import shutil  # We'll need this to remove subdirs
import sys
from concurrent.futures import ThreadPoolExecutor

import openpyxl
import requests
from langdetect import LangDetectException, detect
from openpyxl import Workbook
from requests.exceptions import RequestException, SSLError

# Set up a logger with a dynamic file name
script_name = os.path.splitext(os.path.basename(sys.argv[0]))[0]  # e.g. my_script
time_string = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
log_filename = f"{script_name}_{time_string}.log"

logging.basicConfig(
    filename=log_filename,
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
logger = logging.getLogger(__name__)

API_BASE_URL = "https://api.unpaywall.org/v2/search"
EMAIL = "xlct43@durham.ac.uk"  # Your email for Unpaywall API
EXCEL_DIR = "Excel-Files"
DOWNLOAD_DIR = "Downloaded-Unpaywall"


def clear_directory(path):
    """
    Remove all contents (files/subdirs) inside 'path', but leave the directory itself.
    """
    if os.path.isdir(path):
        for filename in os.listdir(path):
            full_path = os.path.join(path, filename)
            if os.path.isfile(full_path) or os.path.islink(full_path):
                os.remove(full_path)
            elif os.path.isdir(full_path):
                shutil.rmtree(full_path)


# 1) Make directories if they don't exist
os.makedirs(EXCEL_DIR, exist_ok=True)
os.makedirs(DOWNLOAD_DIR, exist_ok=True)

# 2) Clear them so we start fresh
clear_directory(EXCEL_DIR)
clear_directory(DOWNLOAD_DIR)


def single_request(url, params):
    """
    Makes a single HTTP GET request for the given URL and params.
    Returns the JSON data if successful, or None if it fails once.
    """
    try:
        response = requests.get(url, params=params, timeout=30)
        response.raise_for_status()
        return response.json()
    except RequestException as e:
        logger.error(f"[ERROR] Request failed: {e}")
        return None


def sanitize_filename(name):
    """Replace unsafe filename characters with underscores."""
    return re.sub(r"[^a-zA-Z0-9_.-]", "_", name)


def generate_filenames(query_keywords):
    """
    Given a list of query keywords, produce:
    - A safe Excel filename
    - A safe directory name for downloaded PDFs
    """
    safe_query = sanitize_filename("_".join(query_keywords))
    excel_filename = f"{safe_query}.xlsx"
    download_directory = safe_query
    return excel_filename, download_directory


def search_unpaywall_by_title(
    keywords,
    email,
    is_oa=None,
    max_pages=5,
    output_xlsx="unpaywall_results.xlsx",
    wait_s=1.0,
):
    """
    Searches the Unpaywall API and stores metadata in three sheets:
      - 'All'        : All articles retrieved
      - 'Accessible' : Will later store non-closed papers
      - 'Downloaded' : Will later store info on successfully downloaded papers

    *** Modified to skip non-English articles using langdetect. ***
    *** Single-attempt request instead of multiple retries. ***
    """
    output_xlsx = os.path.join(EXCEL_DIR, output_xlsx)

    wb = Workbook()
    sheet_all = wb.active
    sheet_all.title = "All"
    sheet_all.append(
        [
            "doi",
            "title",
            "publisher",
            "year",
            "oa_status",
            "best_oa_location_url",
            "journal_name",
        ]
    )

    wb.create_sheet("Accessible")
    wb.create_sheet("Downloaded")
    wb.save(output_xlsx)

    for page_num in range(1, max_pages + 1):
        params = {"query": keywords, "email": email, "page": page_num}
        if is_oa is not None:
            params["is_oa"] = str(is_oa).lower()

        logger.info(f"[INFO] Fetching page {page_num} for query='{keywords}'...")

        data = single_request(API_BASE_URL, params)
        if not data or "results" not in data:
            logger.info(f"[INFO] No more articles found. Stopping at page {page_num}.")
            break

        wb = openpyxl.load_workbook(output_xlsx)
        sheet_all = wb["All"]

        for result in data["results"]:
            doi_obj = result.get("response", {})
            title = doi_obj.get("title") or ""
            if not title.strip():
                continue

            # Language detection for English only
            try:
                lang = detect(title)
            except LangDetectException:
                continue

            if lang != "en":
                continue

            best_oa_loc = doi_obj.get("best_oa_location") or {}
            best_oa_url = best_oa_loc.get("url_for_pdf") or best_oa_loc.get("url")

            row_data = [
                doi_obj.get("doi"),
                title,
                doi_obj.get("publisher"),
                doi_obj.get("year"),
                doi_obj.get("oa_status"),
                best_oa_url,
                doi_obj.get("journal_name"),
            ]
            sheet_all.append(row_data)

        wb.save(output_xlsx)

        # time.sleep(wait_s)

    logger.info(f"[DONE] Results saved to '{output_xlsx}'.")


def filter_oa_status_not_closed(excel_file):
    """
    Reads the 'All' sheet, keeps rows with oa_status != 'closed' in 'Accessible'.
    """
    excel_path = os.path.join(EXCEL_DIR, excel_file)
    wb = openpyxl.load_workbook(excel_path)

    if "All" not in wb.sheetnames:
        logger.warning("[WARNING] 'All' sheet not found; cannot filter.")
        return

    sheet_all = wb["All"]
    if "Accessible" in wb.sheetnames:
        del wb["Accessible"]
    sheet_accessible = wb.create_sheet("Accessible")

    sheet_accessible.append([cell.value for cell in sheet_all[1]])
    oa_status_index = 4
    for row in sheet_all.iter_rows(min_row=2, values_only=True):
        if row[oa_status_index] != "closed":
            sheet_accessible.append(row)

    wb.save(excel_path)
    logger.info(f"[INFO] Filtered data saved in 'Accessible' sheet of '{excel_file}'.")


def download_paper(row, directory, sheet_downloaded):
    """
    Attempts to download the paper from the best_oa_location_url.
    If successful, appends the row's metadata to the 'Downloaded' sheet.

    *** PRINTS the server's status code and Content-Type to help debug. ***
    """
    doi, url = row[0], row[5]
    if not url:
        return f"[INFO] No URL for DOI '{doi}', skipping."

    safe_doi = sanitize_filename(doi or "UNKNOWN_DOI")
    local_pdf_path = os.path.join(directory, f"{safe_doi}.pdf")

    try:
        response = requests.get(url, stream=True, timeout=30)

        status_code = response.status_code
        content_type = response.headers.get("Content-Type", "").lower()

        disposition = response.headers.get("Content-Disposition", "")
        is_pdf = content_type.startswith("application/pdf") or ".pdf" in disposition

        if response.status_code == 200 and is_pdf:
            with open(local_pdf_path, "wb") as f:
                for chunk in response.iter_content(chunk_size=8192):
                    f.write(chunk)

            sheet_downloaded.append(row)
            return f"[INFO] Downloaded: {local_pdf_path}"
        else:
            return f"[WARNING] Not a PDF or download failed. URL={url}"

    except SSLError as ssl_err:
        return f"[ERROR] SSL error for DOI '{doi}': {ssl_err}"

    except RequestException as e:
        return f"[ERROR] Download failed for DOI '{doi}': {e}"


def download_papers_from_accessible(excel_file, directory_name):
    """
    Reads the 'Accessible' sheet, downloads PDFs to a subdirectory in DOWNLOAD_DIR,
    and logs each successful download in the 'Downloaded' sheet.
    """
    excel_path = os.path.join(EXCEL_DIR, excel_file)
    wb = openpyxl.load_workbook(excel_path)

    if "Accessible" not in wb.sheetnames:
        logger.warning(
            "[WARNING] 'Accessible' sheet not found. No downloads attempted."
        )
        return

    if "Downloaded" not in wb.sheetnames:
        wb.create_sheet("Downloaded")

    sheet_accessible = wb["Accessible"]
    sheet_downloaded = wb["Downloaded"]
    directory = os.path.join(DOWNLOAD_DIR, directory_name)
    os.makedirs(directory, exist_ok=True)
    rows = [row for row in sheet_accessible.iter_rows(min_row=2, values_only=True)]

    with ThreadPoolExecutor(max_workers=5) as executor:
        results = executor.map(
            lambda r: download_paper(r, directory, sheet_downloaded), rows
        )

    for result in results:
        logger.info(result)

    wb.save(excel_path)


painters = [
    "caravaggio",
    "vermeer",
    "van gogh",
    "rembrandt",
    "da vinci",
    "raphael",
    "titian",
    "rubens",
    "el greco",
    "bosch",
    "poussin",
    "angelico",
    "monet",
    "renoir",
    "degas",
    "manet",
    "pissarro",
    "cezanne",
    "seurat",
    "gauguin",
    "picasso",
    "matisse",
    "dali",
    "kandinsky",
    "munch",
    "chagall",
    "pollock",
    "rothko",
    "o'keeffe",
    "hopper",
    "basquiat",
    "lichtenstein",
    "warhol",
    "miro",
    "ernst",
    "tapies",
    "escher",
    "klee",
    "gainsborough",
    "constable",
    "whistler",
    "courbet",
    "millet",
    "delacroix",
    "ingres",
    "david",
    "kahlo",
    "orozco",
    "rivera",
    "siqueiros",
    "botero",
    "qi",
    "fan",
    "hokusai",
    "hiroshige",
    "jeong",
    "varma",
    "sher-gil",
    "farshchian",
    "enwonwu",
    "repin",
    "ayvazovsky",
    "murakami",
    "soutine",
    "modigliani",
    "haring",
    "rossetti",
    "bacon",
    "fragonard",
    "bouguereau",
    "corot",
    "lautrec",
    "murillo",
    "sargent",
    "durer",
    "botticelli",
    "xu",
    "huang",
    "lin",
    "zhang",
    "klimt",
    "goya",
    "eakins",
    "turner",
    "sorolla",
    "cassatt",
    "chardin",
    "van dyck",
    "velazquez",
    "zorn",
    "holbein",
    "boucher",
    "hals",
    "redon",
    "matta",
    "ruscha",
    "watteau",
    "verrocchio",
    "magritte",
    "schiele",
    "hockney",
    "de chirico",
    "van eyck",
    "masaccio",
    "perugino",
    "morandi",
    "frankenthaler",
    "foujita",
    "fini",
    "levitan",
    "de kooning",
    "gervex",
    "boudin",
    "grosz",
    "tanguy",
    "balthus",
    "wu",
    "lee",
    "choi",
    "signac",
    "mantegna",
    "bellini",
    "giorgione",
    "freud",
    "masolino",
    "glackens",
    "raeburn",
    "kusama",
    "correggio",
    "burne-jones",
    "millais",
    "moreau",
    "bruegel",
    "ruisdael",
    "lowry",
    "poulakis",
    "shishkin",
    "hirst",
    "shahn",
    "diebenkorn",
    "leger",
    "kentridge",
    "vasari",
    "bronzino",
    "veronese",
    "altdorfer",
    "grunewald",
    "schongauer",
    "tiepolo",
    "guardi",
    "canaletto",
    "morisot",
    "liu",
    "hishida",
    "delaunay",
    "caillebotte",
    "pinturicchio",
    "allori",
    "pontormo",
    "gris",
    "rouault",
    "gorky",
    "reinhardt",
    "louis",
    "shitao",
    "guo",
    "liang",
    "roy",
    "kwon",
    "fouquet",
    "baldung",
    "braque",
    "serov",
    "boccioni",
    "carpaccio",
    "kirchner",
    "segantini",
    "domenichino",
    "peale",
    "bazille",
    "girodet",
    "rousseau",
    "bouts",
    "memling",
    "van der velde",
    "crivelli",
    "moholy-nagy",
    "opie",
    "johns",
    "beuys",
    "delvaux",
    "stella",
    "reynolds",
    "blake",
    "hodler",
    "la tour",
    "lissitzky",
    "kiefer",
    "parrish",
    "benton",
]


for painter in painters:
    query_keywords = [painter, "art"]
    joined_query = " ".join(query_keywords)

    msg = f"[*] Searching with query: {joined_query}"
    print(msg)  # keep this final print statement
    logger.info(msg)

    excel_filename, download_dir_name = generate_filenames(query_keywords)
    search_unpaywall_by_title(
        joined_query,
        EMAIL,
        is_oa=None,
        max_pages=3,
        output_xlsx=excel_filename,
        wait_s=1.0,
    )
    filter_oa_status_not_closed(excel_filename)
    download_papers_from_accessible(excel_filename, directory_name=download_dir_name)
