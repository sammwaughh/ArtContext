#!/usr/bin/env python3
"""
Stage 3 of the pipeline – harvest OpenAlex literature for a set of painters,
store the metadata in Excel and download the relevant PDFs.

CLI
---
python 03_openalex_download.py                 # all painters
python 03_openalex_download.py -r 150          # first 150
python 03_openalex_download.py -r 350:420      # rows 350 – 420 (1-based, inclusive)

Style
-----
• PEP 526 type annotations
• SCREAMING_SNAKE_CASE constants
• Modern Python 3.9+ syntax
"""

from __future__ import annotations

import argparse
import asyncio
import json
import logging
import os
import re
import threading
import time
from logging.handlers import TimedRotatingFileHandler
from pathlib import Path
from typing import Dict, List, Tuple

import httpx
import pandas as pd
import psutil
import requests
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from requests.adapters import HTTPAdapter
from tqdm import tqdm
from urllib3.util.retry import Retry

# ──────────────────────────── configuration ───────────────────────────────────

CONTACT_EMAIL: str = "samjmwaugh@gmail.com"

# ── Excel folders ------------------------------------------------------------
EXCEL_DIR: Path = Path("Excel-Files")  # root for every workbook
WORKS_DIR: Path = EXCEL_DIR / "Painter-Works-Metadata"  # per-painter workbooks
PAINTERS_XLSX: Path = EXCEL_DIR / "painters.xlsx"  # input list
# ── JSON metadata -----------------------------------------------------------
JSON_DIR: Path = Path("Artist-JSONs")  # new – per-painter JSON

# ── Output PDFs --------------------------------------------------------------
# keep them inside the Pipeline package: “…/Pipeline/PDFs/”
PDF_DIR: Path = Path(__file__).resolve().parent / "PDFs"

# ── Logs ---------------------------------------------------------------------
LOG_ROOT: Path = Path("logs")
PAINTER_LOG_DIR: Path = LOG_ROOT / "painter-logs"
LOG_DIRS: Tuple[Path, ...] = (
    LOG_ROOT / "fetch-logs",
    LOG_ROOT / "excel-logs",
    LOG_ROOT / "download-logs",
    LOG_ROOT / "cpu-logs",
)

# Topics centred on the historical, cultural and iconographic interpretation
# of painting (ancient → contemporary).  We fetch any work tagged with at least
# one of the following confirmed OpenAlex topic IDs.
TOPIC_IDS: str = (
    # ── General art-historical / theoretical topics ────────────────────────
    "C52119013|"  # Art History
    "T13922|"  # Historical Art and Culture Studies
    "T12632|"  # Visual Culture and Art Theory
    "T12650|"  # Aesthetic Perception and Analysis
    "C204034006|"  # Art Criticism
    "C501303744|"  # Iconography
    # ── Period-based topics ────────────────────────────────────────────────
    "C554736915|"  # Ancient Art
    "C138634970|"  # Medieval Art
    "T12076|"  # Renaissance and Early Modern Studies
    "C189135316|"  # Modern Art
    "C85363599|"  # Contemporary Art
    # ── Movement-specific topics ───────────────────────────────────────────
    "C32685002|"  # Romanticism
    # ── Regional topics ────────────────────────────────────────────────────
    "C12183850|"  # Indian / Asian Art
    "C2993994385|"  # Islamic Art
    "C64626740"  # African Art
)

DEFAULT_PER_PAGE: int = 200  # OpenAlex page size (API max)
PDF_CHUNK_SIZE: int = 8_192  # bytes
CPU_SAMPLE_SEC: int = 1
MAX_RETRIES: int = 5  # more resilience
BACKOFF_FACTOR: float = 0.5  # 0.5 → 0.5,1,2,4,8 s back-off
SESSION_TIMEOUT_SEC: int = 30  # large pages need a bit more
MAX_WORKERS_FACTOR: int = 3  # ≈ workers = CPUs × factor
MIN_WORKERS: int = 6

START_ROW: int = 1  # default CLI range start (1-based)
END_ROW: int | None = None  # default CLI range end   (inclusive)
DEFAULT_SLEEP_SEC: float = 0.15  # ≈6-7 rps < OpenAlex 10 rps cap

# ── Filtering ----------------------------------------------------------------
# Keep only works whose OpenAlex relevance_score exceeds this value
RELEVANCE_THRESHOLD: float = 1.0

# Only fetch columns we actually use (smaller payloads, faster)
# SELECT_FIELDS: str = (
#     "id,display_name,relevance_score,doi,primary_location,type,"
#     "open_access,locations,best_oa_location"
# )
# need the concepts list to attach topic IDs
SELECT_FIELDS: str = (
    "id,display_name,relevance_score,doi,primary_location,type,"
    "open_access,locations,best_oa_location,concepts"
)

# ──────────────────────────── helpers & utils ─────────────────────────────────


def _make_session() -> requests.Session:
    """Return a shared, retry-enabled HTTP session."""
    retry: Retry = Retry(
        total=MAX_RETRIES,
        backoff_factor=BACKOFF_FACTOR,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["HEAD", "GET", "OPTIONS"],
        raise_on_status=False,
    )
    sess: requests.Session = requests.Session()
    adapter: HTTPAdapter = HTTPAdapter(max_retries=retry)
    sess.mount("https://", adapter)
    sess.mount("http://", adapter)
    sess.headers.update({"User-Agent": f"ArtContext/0.1 (mailto:{CONTACT_EMAIL})"})
    return sess


SESSION: requests.Session = _make_session()


def _setup_logger() -> logging.Logger:
    """One rotating root logger (midnight UTC)."""
    log_dir: Path = Path("logs")
    log_dir.mkdir(exist_ok=True)
    handler = TimedRotatingFileHandler(
        log_dir / "openalex.log", when="midnight", utc=True, backupCount=7
    )
    handler.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
    root = logging.getLogger("openalex")
    root.setLevel(logging.INFO)
    root.addHandler(handler)
    return root


LOGGER: logging.Logger = _setup_logger()

# ─────────────────────────── file-level log helper ───────────────────────────


def _file_logger(path: Path) -> logging.Logger:
    """
    Return a fresh logger that writes *only* to *path*.
    The logger name is the stem of the file so each painter gets its own one.
    """
    logger = logging.getLogger(path.stem)
    logger.setLevel(logging.INFO)
    logger.propagate = False  # don’t double-write via root
    if logger.hasHandlers():
        logger.handlers.clear()
    handler = logging.FileHandler(path)
    handler.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
    logger.addHandler(handler)
    return logger


def _safe_json_parse(value: str | Dict | List) -> Dict | List | None:
    """`json.loads` that survives invalid input."""
    if isinstance(value, (dict, list)):
        return value
    try:
        return json.loads(value)
    except Exception:  # noqa: BLE001
        return None


def _sanitize_filename(name, max_len: int = 120) -> str:  # accept any type
    """
    Return *name* as a safe filename component.

    • Converts non-string input via ``str``.
    • Strips characters that are illegal in most filesystems.
    • Falls back to ``untitled`` when result is empty.
    """
    if not isinstance(name, str):
        try:
            name = str(name)
        except Exception:  # noqa: BLE001
            name = ""

    # Strip filesystem-hostile chars
    cleaned: str = re.sub(r'[\\/*?:"<>|]', "", name).strip()

    # Empty, “nan”, or still non-string → fallback
    if not cleaned or cleaned.lower() == "nan":
        cleaned = "untitled"

    # Avoid overlong paths (reserve room for prefix/suffix)
    if len(cleaned) > max_len:
        cleaned = cleaned[:max_len].rstrip()

    return cleaned


# ──────────────────────────── OpenAlex access ─────────────────────────────────


def fetch_works(
    painter: str,
    per_page: int,
    sleep_sec: float,
) -> List[Dict]:
    """
    Retrieve every OpenAlex work that is
    English, open-access, within TOPIC_IDS and mentions *painter*.
    """
    base_url: str = "https://api.openalex.org/works"
    query_filter: str = f"language:en,is_oa:true,topics.id:{TOPIC_IDS}"
    cursor: str | None = "*"
    works: List[Dict] = []
    page: int = 0

    LOGGER.info("Query start – painter=%s", painter)
    pbar = tqdm(desc=f"{painter} pages", unit="page")

    while cursor:
        page += 1
        pbar.update(1)
        params: Dict[str, str | int | float] = {
            "filter": query_filter,
            # wrap the name in quotes so OpenAlex treats it as an exact phrase
            "search": painter,
            "per_page": per_page,
            "select": SELECT_FIELDS,
            "cursor": cursor,
            "mailto": CONTACT_EMAIL,
        }
        while True:
            resp = SESSION.get(base_url, params=params, timeout=SESSION_TIMEOUT_SEC)
            if resp.status_code != 429:
                break
            retry_after = int(resp.headers.get("Retry-After", "60"))
            LOGGER.warning("429 – sleeping %d s", retry_after)
            time.sleep(retry_after)

        if resp.status_code != 200:  # give up – log & abort
            LOGGER.error("HTTP %s – %s", resp.status_code, resp.text[:200])
            break

        data: Dict = resp.json()
        page_results: List[Dict] = data.get("results", [])
        works.extend(page_results)
        cursor = data.get("meta", {}).get("next_cursor")

        LOGGER.info("%s page %d – %d works", painter, page, len(page_results))
        time.sleep(sleep_sec)

    pbar.close()
    LOGGER.info("%s finished – pages=%d  works=%d", painter, page, len(works))
    # Early relevance filter
    return [w for w in works if w.get("relevance_score", 0) > RELEVANCE_THRESHOLD]


# ───────────────────────── JSON metadata writer ──────────────────────────────
def _concept_ids(concepts_field) -> List[str]:
    """Extract bare OpenAlex concept/topic IDs from the `concepts` array."""
    ids: List[str] = []
    concepts = _safe_json_parse(concepts_field)
    if not concepts:
        return ids
    for c in concepts:
        if isinstance(c, dict) and "id" in c:
            # keep just the final path segment, e.g. “…/C554736915”
            ids.append(c["id"].rsplit("/", 1)[-1])
    return ids


def write_json_metadata(painter: str, works: List[Dict]) -> Path:
    """
    Dump a per-painter JSON file with the essential metadata
    (id, title, relevance_score, tags, PDF links …).
    """
    rows: List[Dict] = []
    for w in works:
        best, backup = _best_and_backup(pd.Series(w))
        rows.append(
            {
                "id": w.get("id", ""),
                "title": w.get("display_name", ""),
                "relevance_score": w.get("relevance_score", 0.0),
                "tags": _concept_ids(w.get("concepts")),
                "doi": w.get("doi"),
                "type": w.get("type"),
                "best_pdf": best,
                "backup_pdf": backup,
            }
        )

    JSON_DIR.mkdir(exist_ok=True)
    dest: Path = JSON_DIR / f"{painter.lower()}.json"
    dest.write_text(json.dumps(rows, ensure_ascii=False, indent=2))
    LOGGER.info("%s JSON metadata saved (%d works)", painter, len(rows))
    return dest


# ─────────────────────────── Excel creation ─────────────────────────────────────

MAIN_COLS: Tuple[str, ...] = (
    "title",  # we rename display_name → title below
    "relevance_score",
    "id",
    "doi",
    "primary_location",
    "type",
    "open_access",
    "locations",
    "best_oa_location",
)


def _get_candidate_links(row: pd.Series) -> List[str]:
    """Collect all distinct OA/PDF URLs from the various location blobs."""
    candidates: List[str] = []
    best_oa = _safe_json_parse(row["best_oa_location"])
    if isinstance(best_oa, dict):
        candidates += [best_oa.get("pdf_url"), best_oa.get("landing_page_url")]

    oa = _safe_json_parse(row["open_access"])
    if isinstance(oa, dict):
        candidates.append(oa.get("oa_url"))

    primary = _safe_json_parse(row["primary_location"])
    if isinstance(primary, dict):
        candidates += [primary.get("pdf_url"), primary.get("landing_page_url")]

    for loc in _safe_json_parse(row["locations"]) or []:
        if isinstance(loc, dict):
            candidates += [loc.get("pdf_url"), loc.get("landing_page_url")]

    return [c for c in dict.fromkeys(candidates) if c]  # dedupe & drop None


def _best_and_backup(row: pd.Series) -> Tuple[str, str]:
    """Return “best” PDF link plus a backup (may be empty strings)."""
    for_best: List[str] = _get_candidate_links(row)
    best: str = ""
    backup: str = ""
    for link in for_best:
        if ".pdf" in link.lower():
            best = link
            break
    if not best and for_best:
        best = for_best[0]
    for link in for_best:
        if link != best:
            backup = link
            break
    return best or "", backup or ""


def create_excel(painter: str, works: List[Dict]) -> Path:
    """Write three sheets (Main, Filtered, Downloadable) and return the file path."""
    # Build DataFrame and normalise column names
    df_main: pd.DataFrame = pd.DataFrame(works)
    if "display_name" in df_main.columns and "title" not in df_main.columns:
        df_main = df_main.rename(columns={"display_name": "title"})

    # Ensure every expected column exists so `.loc[:, MAIN_COLS]` never fails
    for col in MAIN_COLS:
        if col not in df_main.columns:
            df_main[col] = pd.NA

    df_filtered: pd.DataFrame = df_main.loc[:, MAIN_COLS].copy()

    rows_downloadable: List[Dict[str, str | float]] = []
    for _, row in df_filtered.iterrows():
        best, backup = _best_and_backup(row)
        rows_downloadable.append(
            {
                "Title": row.get("title", ""),
                "Relevance Score": row.get("relevance_score", 0.0),
                "Best Link": best,
                "Backup Link": backup,
                "OpenAlex ID": row.get("id", ""),
                "Type": row.get("type", ""),
            }
        )
    df_download: pd.DataFrame = pd.DataFrame(rows_downloadable)

    dest: Path = WORKS_DIR / f"{painter.lower()}_works.xlsx"
    with pd.ExcelWriter(dest, engine="openpyxl") as writer:
        df_main.to_excel(writer, index=False, sheet_name="Main")
        df_filtered.to_excel(writer, index=False, sheet_name="Filtered")
        df_download.to_excel(writer, index=False, sheet_name="Downloadable")

    # autosize columns (rough)
    wb = load_workbook(dest)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for column in ws.columns:
            width: int = max(len(str(cell.value)) for cell in column if cell.value) + 2
            ws.column_dimensions[get_column_letter(column[0].column)].width = min(
                width, 50
            )
    wb.save(dest)
    LOGGER.info("%s workbook saved (%d rows)", painter, len(df_main))
    return dest


# ───────────────────────── PDF download (async) ────────────────────────────


async def _grab_pdf(
    idx: int,
    title: str,
    best: str,
    backup: str,
    sem: asyncio.Semaphore,
    out_dir: Path,
) -> Tuple[str, float] | None:
    """Async download with automatic Retry-After handling."""
    filename = f"{idx}-{_sanitize_filename(title)}.pdf"
    path = out_dir / filename
    start = time.perf_counter()

    async with sem:
        async with httpx.AsyncClient(
            timeout=SESSION_TIMEOUT_SEC,
            limits=httpx.Limits(max_keepalive_connections=20),
        ) as client:
            for url in (best, backup):
                if not url:
                    continue
                try:
                    r = await client.get(url, follow_redirects=True)
                    if r.status_code == 429:
                        retry_after = int(r.headers.get("Retry-After", "60"))
                        await asyncio.sleep(retry_after)
                        continue
                    if r.status_code != 200:
                        LOGGER.warning(
                            "Row %d – HTTP %s on %s", idx, r.status_code, url
                        )
                        continue
                    path.write_bytes(r.content)
                    return filename, round(time.perf_counter() - start, 3)
                except Exception as exc:  # noqa: BLE001
                    LOGGER.error("Row %d – %s on %s", idx, exc, url)
    return None


def _monitor_cpu(stop_evt: threading.Event, clog: logging.Logger) -> None:
    """Log instantaneous and average CPU usage until *stop_evt* is set."""
    samples: List[float] = []
    while not stop_evt.wait(CPU_SAMPLE_SEC):
        usage: float = psutil.cpu_percent()
        samples.append(usage)
        clog.info("CPU %.1f%%", usage)
    if samples:
        clog.info("Average CPU %.2f%%", sum(samples) / len(samples))


async def download_pdfs(
    excel_file: Path,
    painter: str,
    max_workers: int,
    dlog: logging.Logger,
    clog: logging.Logger,
) -> List[Tuple[str, float]]:
    """
    Parallel PDF download (relevance_score > 1).
    Returns list of (pdf_name, seconds).
    """
    df: pd.DataFrame = pd.read_excel(
        excel_file, sheet_name="Downloadable", engine="openpyxl"
    )
    df = df.drop_duplicates(subset=["Title"], keep="first")

    out_dir: Path = PDF_DIR / painter.lower()
    out_dir.mkdir(parents=True, exist_ok=True)

    stop_evt = threading.Event()
    threading.Thread(target=_monitor_cpu, args=(stop_evt, clog), daemon=True).start()

    sem = asyncio.Semaphore(max_workers)
    tasks: List[asyncio.Task] = []
    # enumerate → (idx, row)  — avoids the “too many values to unpack” crash
    for idx, row in enumerate(df.itertuples(index=False, name=None), start=1):
        title, best, backup = row[0], row[2], row[3]  # relies on sheet layout
        # Be resilient if someone re-orders columns
        if isinstance(row, tuple) and hasattr(df, "columns"):
            mapping = dict(zip(df.columns, row))
            title = mapping.get("Title", title)
            best = mapping.get("Best Link", best)
            backup = mapping.get("Backup Link", backup)

        tasks.append(_grab_pdf(idx, title, best, backup, sem, out_dir))
    results: List[Tuple[str, float]] = []
    for coro in tqdm(
        asyncio.as_completed(tasks), total=len(tasks), desc=f"{painter} PDFs"
    ):
        res = await coro
        if res:
            results.append(res)

    stop_evt.set()

    failed = {row[0] for row in df.itertuples(index=False)} - {r[0] for r in results}
    if failed:
        (out_dir / "failed_downloads.txt").write_text("\n".join(sorted(failed)))
    return results


def append_download_times(excel_file: Path, times: List[Tuple[str, float]]) -> None:
    """Append a ‘Download Times’ sheet to *excel_file* (overwrites if exists)."""
    if not times:
        return
    df: pd.DataFrame = pd.DataFrame(times, columns=["PDF Name", "Seconds"])
    df = df.sort_values("Seconds", ascending=False).reset_index(drop=True)

    wb = load_workbook(excel_file)
    if "Download Times" in wb.sheetnames:
        wb.remove(wb["Download Times"])
    ws = wb.create_sheet("Download Times")
    ws.append(["PDF Name", "Seconds"])
    for name, sec in df.itertuples(index=False):
        ws.append([name, sec])
    for column in ws.columns:
        width = max(len(str(c.value)) for c in column if c.value) + 2
        ws.column_dimensions[get_column_letter(column[0].column)].width = width
    wb.save(excel_file)


# ───────────────────────── painter wrapper ────────────────────────────────────


def process_painter(
    painter: str,
    max_workers: int,
    per_page: int,
    sleep_sec: float,
) -> None:
    """One-stop processing for a single painter."""
    ts: str = time.strftime("%Y%m%d-%H%M%S")
    dlog = _file_logger(PAINTER_LOG_DIR / f"{painter}-{ts}-download.log")
    clog = _file_logger(PAINTER_LOG_DIR / f"{painter}-{ts}-cpu.log")

    # fetch & store
    start_time: float = time.perf_counter()
    works = fetch_works(painter, per_page=per_page, sleep_sec=sleep_sec)
    if not works:
        print("  no works found – skipped")
        return

    excel_file: Path = create_excel(painter, works)
    write_json_metadata(painter, works)  # new JSON output

    times = asyncio.run(download_pdfs(excel_file, painter, max_workers, dlog, clog))
    append_download_times(excel_file, times)

    elapsed: float = time.perf_counter() - start_time
    print(f"  done in {elapsed:.1f} s  ({len(times)} PDFs)")


# ─────────────────────────── main entry point ────────────────────────────────


def _parse_range(text: str) -> Tuple[int, int | None]:
    """
    Convert “N” or “A:B” into (start, end)  – both 1-based, *end* inclusive.
    Validation is minimal; ValueError propagates to argparse.
    """
    if ":" in text:
        a, b = (int(x) for x in text.split(":", 1))
        return a, b
    return 1, int(text)


def main() -> None:
    """CLI front-door."""
    parser = argparse.ArgumentParser(
        description="Download OpenAlex PDFs for painters "
        "(page size, sleep time and workers are taken from the constants "
        "defined at the top of the file)."
    )
    parser.add_argument(
        "-r",
        "--rows",
        metavar="N | A:B",
        help="Rows in painters.xlsx (1-based).  E.g. 100  or  350:420.",
    )
    parser.add_argument(
        "-w",
        "--workers",
        type=int,
        help="Number of concurrent workers (default: auto-detect)",
    )
    parser.add_argument(
        "-p",
        "--per-page",
        type=int,
        default=DEFAULT_PER_PAGE,
        help="Results per page for OpenAlex API (default: %(default)d)",
    )
    parser.add_argument(
        "-s",
        "--sleep",
        type=float,
        default=DEFAULT_SLEEP_SEC,
        help="Sleep time between requests (default: %(default).1f s)",
    )
    args = parser.parse_args()

    start_row, end_row = (
        (START_ROW, END_ROW) if not args.rows else _parse_range(args.rows)
    )

    # ensure directories
    EXCEL_DIR.mkdir(exist_ok=True)
    WORKS_DIR.mkdir(exist_ok=True)
    PDF_DIR.mkdir(exist_ok=True)
    JSON_DIR.mkdir(exist_ok=True)
    for d in (*LOG_DIRS, PAINTER_LOG_DIR):
        d.mkdir(parents=True, exist_ok=True)

    df_painters: pd.DataFrame = pd.read_excel(PAINTERS_XLSX)
    subset: pd.DataFrame = df_painters.iloc[
        start_row - 1 : end_row
    ]  # end_row may be None

    max_workers: int = max(MIN_WORKERS, os.cpu_count() * MAX_WORKERS_FACTOR)
    if args.workers:
        max_workers = max(1, args.workers)

    for i, row in enumerate(subset.itertuples(index=False), start=start_row):
        # Get the painter’s query string
        # (supports either “Query String” or single-column layout)
        painter: str = (
            row[df_painters.columns.get_loc("Query String")]
            if "Query String" in df_painters.columns
            else row[0]
        )
        # Immediate, informative progress line
        print(f"Row {i:>4} – {painter}", flush=True)
        process_painter(
            painter,
            max_workers=max_workers,
            per_page=args.per_page,
            sleep_sec=args.sleep,
        )


if __name__ == "__main__":
    main()
