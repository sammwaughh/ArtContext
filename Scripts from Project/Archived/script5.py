import os
import re
import sys
import time
from concurrent.futures import ThreadPoolExecutor

import openpyxl
import requests
from openpyxl import Workbook

API_BASE_URL = "https://api.unpaywall.org/v2/search"
EMAIL = "xlct43@durham.ac.uk"  # Your email for Unpaywall API
EXCEL_DIR = "Excel-Files"
DOWNLOAD_DIR = "Downloaded-Unpaywall"

os.makedirs(EXCEL_DIR, exist_ok=True)
os.makedirs(DOWNLOAD_DIR, exist_ok=True)


def retry_request(url, params, retries=3, wait_s=2.0):
    """Retries HTTP GET requests with exponential backoff on failure."""
    for attempt in range(retries):
        try:
            response = requests.get(url, params=params, timeout=30)
            response.raise_for_status()
            return response.json()
        except requests.exceptions.RequestException as e:
            print(f"[ERROR] Attempt {attempt + 1}: {e}")
            time.sleep(wait_s * (2**attempt))  # Exponential backoff
    return None  # Return None if all attempts fail


def sanitize_filename(name):
    """Replace unsafe filename characters with underscores."""
    return re.sub(r"[^a-zA-Z0-9_.-]", "_", name)


def generate_filenames(query_keywords):
    """
    Given a list of query keywords, produce:
    - A safe Excel filename
    - A safe directory name for downloaded PDFs
    """
    safe_query = sanitize_filename("_".join(query_keywords))
    excel_filename = f"{safe_query}.xlsx"
    download_directory = safe_query
    return excel_filename, download_directory


def search_unpaywall_by_title(
    keywords,
    email,
    is_oa=None,
    max_pages=5,
    output_xlsx="unpaywall_results.xlsx",
    wait_s=1.0,
):
    """
    Searches the Unpaywall API and stores metadata in three sheets:
      - 'All'        : All articles retrieved
      - 'Accessible' : Will later store non-closed papers
      - 'Downloaded' : Will later store info on successfully downloaded papers
    """
    # Ensure the Excel file is created within EXCEL_DIR
    output_xlsx = os.path.join(EXCEL_DIR, output_xlsx)

    # Create a new workbook with our three sheets
    wb = Workbook()
    sheet_all = wb.active
    sheet_all.title = "All"
    sheet_all.append(
        [
            "doi",
            "title",
            "publisher",
            "year",
            "oa_status",
            "best_oa_location_url",
            "journal_name",
        ]
    )

    # Create 'Accessible' and 'Downloaded' sheets
    wb.create_sheet("Accessible")
    wb.create_sheet("Downloaded")

    # Save after creating sheets
    wb.save(output_xlsx)

    # Fetch articles from the API
    for page_num in range(1, max_pages + 1):
        params = {"query": keywords, "email": email, "page": page_num}
        if is_oa is not None:
            params["is_oa"] = str(is_oa).lower()

        print(f"[INFO] Fetching page {page_num}...")
        data = retry_request(API_BASE_URL, params)
        if not data or "results" not in data:
            print(f"[INFO] No more articles found. Stopping at page {page_num}.")
            break

        # Append each result to 'All'
        for result in data["results"]:
            doi_obj = result.get("response", {})
            best_oa_loc = doi_obj.get("best_oa_location") or {}
            best_oa_url = best_oa_loc.get("url_for_pdf") or best_oa_loc.get("url")

            row_data = [
                doi_obj.get("doi"),
                doi_obj.get("title"),
                doi_obj.get("publisher"),
                doi_obj.get("year"),
                doi_obj.get("oa_status"),
                best_oa_url,
                doi_obj.get("journal_name"),
            ]
            sheet_all.append(row_data)

        # Be polite to the API
        time.sleep(wait_s)

    # Final save
    wb.save(output_xlsx)
    print(f"[DONE] Results saved to '{output_xlsx}'.")


def filter_oa_status_not_closed(excel_file):
    """
    Reads the 'All' sheet, keeps rows with oa_status != 'closed' in 'Accessible'.
    """
    # Make sure we read from the correct directory
    excel_path = os.path.join(EXCEL_DIR, excel_file)
    wb = openpyxl.load_workbook(excel_path)

    if "All" not in wb.sheetnames:
        print("[WARNING] 'All' sheet not found; cannot filter.")
        return

    sheet_all = wb["All"]
    # Re-create 'Accessible' each time
    if "Accessible" in wb.sheetnames:
        del wb["Accessible"]
    sheet_accessible = wb.create_sheet("Accessible")

    # Copy header from 'All'
    sheet_accessible.append([cell.value for cell in sheet_all[1]])

    # 'oa_status' is in column index 4
    oa_status_index = 4
    for row in sheet_all.iter_rows(min_row=2, values_only=True):
        if row[oa_status_index] != "closed":
            sheet_accessible.append(row)

    wb.save(excel_path)
    print(f"[INFO] Filtered data saved in 'Accessible' sheet of '{excel_file}'.")


def download_paper(row, directory, sheet_downloaded):
    """
    Attempts to download the paper from the best_oa_location_url.
    If successful, appends the row's metadata to the 'Downloaded' sheet.
    """
    doi, url = row[0], row[5]
    if not url:
        return f"[INFO] No URL for DOI '{doi}', skipping."

    safe_doi = sanitize_filename(doi or "UNKNOWN_DOI")
    local_pdf_path = os.path.join(directory, f"{safe_doi}.pdf")

    try:
        response = requests.get(url, stream=True, timeout=30)
        content_type = response.headers.get("Content-Type", "").lower()
        disposition = response.headers.get("Content-Disposition", "")
        is_pdf = content_type.startswith("application/pdf") or ".pdf" in disposition

        if response.status_code == 200 and is_pdf:
            with open(local_pdf_path, "wb") as f:
                for chunk in response.iter_content(chunk_size=8192):
                    f.write(chunk)

            # Append row to 'Downloaded' if successful
            sheet_downloaded.append(row)
            return f"[INFO] Downloaded: {local_pdf_path}"
        else:
            return f"[WARNING] Not a PDF or download failed. URL={url}"
    except requests.exceptions.RequestException as e:
        return f"[ERROR] Download failed for DOI '{doi}': {e}"


def download_papers_from_accessible(excel_file, directory_name):
    """
    Reads the 'Accessible' sheet, downloads PDFs to a subdirectory in DOWNLOAD_DIR,
    and logs each successful download in the 'Downloaded' sheet.
    """
    excel_path = os.path.join(EXCEL_DIR, excel_file)
    wb = openpyxl.load_workbook(excel_path)

    # Ensure the 'Accessible' sheet exists
    if "Accessible" not in wb.sheetnames:
        print("[WARNING] 'Accessible' sheet not found. No downloads attempted.")
        return

    # Ensure we have a 'Downloaded' sheet
    if "Downloaded" not in wb.sheetnames:
        wb.create_sheet("Downloaded")

    sheet_accessible = wb["Accessible"]
    sheet_downloaded = wb["Downloaded"]

    # Directory for this query under DOWNLOAD_DIR
    directory = os.path.join(DOWNLOAD_DIR, directory_name)
    os.makedirs(directory, exist_ok=True)

    # Read all data from 'Accessible'
    rows = [row for row in sheet_accessible.iter_rows(min_row=2, values_only=True)]

    # Download in parallel using ThreadPoolExecutor
    with ThreadPoolExecutor(max_workers=5) as executor:
        results = executor.map(
            lambda r: download_paper(r, directory, sheet_downloaded), rows
        )

    # Print results
    for result in results:
        print(result)

    # Save the updated workbook
    wb.save(excel_path)


# Main: parse command-line arguments, generate filenames, call pipeline
if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python script5.py <search terms>")
        sys.exit(1)

    # Convert arguments into a single query string and a list for filenames
    query_keywords = sys.argv[1:]  # e.g. ["Caravaggio", "paintings"]
    joined_query = " ".join(query_keywords)  # e.g. "Caravaggio paintings"

    # Build Excel filename and download directory from query
    excel_filename, download_dir_name = generate_filenames(query_keywords)

    # 1. Fetch from Unpaywall, storing everything in 'All'
    search_unpaywall_by_title(
        joined_query,
        EMAIL,
        is_oa=None,
        max_pages=3,
        output_xlsx=excel_filename,
        wait_s=1.0,
    )

    # 2. Filter to create/update 'Accessible'
    filter_oa_status_not_closed(excel_filename)

    # 3. Download from 'Accessible' and log to 'Downloaded'
    download_papers_from_accessible(excel_filename, directory_name=download_dir_name)
