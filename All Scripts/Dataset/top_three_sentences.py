import os
import nltk
import numpy as np
import pandas as pd
import torch
import time
from openpyxl import load_workbook
from sentence_transformers import SentenceTransformer, util
from tqdm import tqdm

def find_matching_paintings(paintings_range, painter_name):
    # Try an exact case-insensitive match first.
    mask_exact = paintings_range["Creator"].str.lower() == painter_name.lower()
    if mask_exact.sum() > 0:
        print(f"Found {mask_exact.sum()} paintings with exact match for '{painter_name}'.")
        return paintings_range[mask_exact]
    
    # Try matching by first name.
    first_name = painter_name.split()[0]
    mask_first = paintings_range["Creator"].str.lower().str.contains(first_name.lower(), na=False)
    if mask_first.sum() > 0:
        print(f"No exact match found. Found {mask_first.sum()} paintings with first name '{first_name}'.")
        return paintings_range[mask_first]
    
    # Try matching by last name.
    last_name = painter_name.split()[-1]
    mask_last = paintings_range["Creator"].str.lower().str.contains(last_name.lower(), na=False)
    if mask_last.sum() > 0:
        print(f"No exact or first name match found. Found {mask_last.sum()} paintings with last name '{last_name}'.")
        return paintings_range[mask_last]
    
    print(f"No paintings found for variations of '{painter_name}'.")
    return pd.DataFrame()

def main():
    nltk.download('punkt')
    nltk.download('punkt_tab')
    
    # Start timer.
    start_time = time.perf_counter()
    
    # Load paintings.xlsx and painters.xlsx
    paintings_df = pd.read_excel("paintings.xlsx")
    painters_df = pd.read_excel("painters.xlsx")
    cache_dir = "cache"
    
    # Specify an inclusive range of paintings (rows in paintings.xlsx) to process (0-indexed)
    painting_start_index = 0
    painting_end_index = 40000  # e.g., process rows 0 through 30000 inclusive
    
    # Specify an inclusive range of painters (rows in painters.xlsx) to process (0-indexed)
    painter_start_index = 0
    painter_end_index = 30   # e.g., process painters rows 0 through 10 inclusive
    
    # Restrict to the specified subsets.
    paintings_range = paintings_df.iloc[painting_start_index:painting_end_index+1]
    painters_subset = painters_df.iloc[painter_start_index:painter_end_index+1]
    
    # Load the existing workbook with openpyxl (to preserve formatting)
    wb = load_workbook("paintings.xlsx")
    ws = wb.active

    # Build a dictionary mapping header names to their column numbers (from header row 1)
    header_map = {}
    for cell in ws[1]:
        if cell.value is not None:
            header_map[cell.value] = cell.column

    device = "mps" if torch.backends.mps.is_available() else "cpu"
    print(f"Using device: {device}")
    print("Loading Sentence-BERT model...")
    model = SentenceTransformer('paraphrase-MiniLM-L6-v2', device=device)
    
    total_processed = 0

    # Iterate over each painter in the specified painters range.
    for _, painter_row in painters_subset.iterrows():
        painter_name = painter_row["Artist"]
        artist_dir_name = painter_row["Query String"]
        print(f"\nProcessing painter: {painter_name} with directory mapping: {artist_dir_name}")
        
        cache_file = os.path.join(cache_dir, f"{artist_dir_name}_cache.pt")
        if not os.path.exists(cache_file):
            print(f"No cache file for painter {painter_name} at {cache_file}. Skipping.")
            continue
        print(f"Loading cache from {cache_file}...")
        data = torch.load(cache_file)
        candidate_contexts_all = data["candidate_contexts"]
        candidate_sentences_all = data["candidate_sentences"]
        candidate_embeddings = data["candidate_embeddings"]
        # Ensure candidate embeddings are on the same device as the query.
        candidate_embeddings = candidate_embeddings.to(device)
        
        # Find matching paintings in the specified range using variations.
        matching_paintings = find_matching_paintings(paintings_range, painter_name)
        if matching_paintings.empty:
            continue
        
        # Process each matching painting.
        for idx, row in matching_paintings.iterrows():
            try:
                painting_title = row["Title"]
                artist = row["Creator"]
            except IndexError:
                print(f"Row index {idx} out of range in paintings.xlsx.")
                continue

            query = f"Describe the visual content of '{painting_title}' by {artist}."
            query_embedding = model.encode(query, convert_to_tensor=True)
            
            # Compute cosine similarity in chunks with progress feedback.
            chunk_size = 10000
            cosine_scores_list = []
            num_candidates = candidate_embeddings.shape[0]
            
            for i in tqdm(range(0, num_candidates, chunk_size), desc="Cosine similarity", leave=False):
                chunk = candidate_embeddings[i:i+chunk_size]
                scores_chunk = util.cos_sim(query_embedding, chunk)[0]
                cosine_scores_list.append(scores_chunk.cpu().numpy())
            cosine_scores_np = np.concatenate(cosine_scores_list)
            top_indices = np.argsort(cosine_scores_np)[-3:][::-1]
            best_sentences = [candidate_sentences_all[i] for i in top_indices]
            
            # Update the corresponding row in the workbook.
            # Excel rows are 1-indexed; header is in row 1, so data rows start at row index+2.
            excel_row = idx + 2
            ws.cell(row=excel_row, column=header_map["Sentence 1"], value=best_sentences[0])
            ws.cell(row=excel_row, column=header_map["Sentence 2"], value=best_sentences[1])
            ws.cell(row=excel_row, column=header_map["Sentence 3"], value=best_sentences[2])
            
            total_processed += 1

    wb.save("paintings.xlsx")
    elapsed_time = time.perf_counter() - start_time
    elapsed_rounded = round(elapsed_time, 1)
    rate = round(elapsed_time / total_processed, 1) if total_processed > 0 else float('inf')
    
    print("\nUpdated paintings.xlsx with the best sentences for all processed paintings.")
    print(f"Processed {total_processed} paintings in {elapsed_rounded} seconds. Rate: {rate} seconds per painting.")

if __name__ == "__main__":
    main()
