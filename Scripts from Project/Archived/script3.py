import os
import time

import openpyxl
import requests
from openpyxl import Workbook

API_BASE_URL = "https://api.unpaywall.org/v2/search"
EMAIL = (
    "xlct43@durham.ac.uk"  # Replace with the email you use for Unpaywall API requests
)


def search_unpaywall_by_title(
    keywords,
    email,
    is_oa=None,
    max_pages=5,
    output_xlsx="unpaywall_results.xlsx",
    wait_s=1.0,
):
    """
    Query the Unpaywall search endpoint with the given keywords, fetch results,
    and save them to an Excel file (first sheet named 'All').

    Args:
        keywords (str): Space-separated keywords (e.g. "Titian paintings")
                        or a more complex query using quoted text, OR, etc.
        email (str)   : Your email parameter required by Unpaywall.
        is_oa (bool)  : If True, only retrieve OA articles;
                        if False, only non-OA; if None, any article.
        max_pages (int): Max pages to retrieve (50 results per page).
        output_xlsx (str): Path for Excel output.
        wait_s (float): Seconds to wait between page requests to avoid
                        flooding the API.
    """

    # Create a new Excel workbook and get the active sheet
    wb = Workbook()
    sheet_all = wb.active
    sheet_all.title = "All"

    # Write header row
    sheet_all.append(
        [
            "doi",
            "title",
            "publisher",
            "year",
            "oa_status",
            "best_oa_location_url",
            "journal_name",
        ]
    )

    # For each page, call the Unpaywall search endpoint
    for page_num in range(1, max_pages + 1):
        params = {"query": keywords, "email": email, "page": page_num}
        if is_oa is True:
            params["is_oa"] = "true"
        elif is_oa is False:
            params["is_oa"] = "false"

        print(f"[INFO] Fetching page {page_num} with params={params}")

        # Make the request
        response = requests.get(API_BASE_URL, params=params)
        response.raise_for_status()  # raise exception if the request failed
        data = response.json()

        # Extract results from JSON
        results = data.get("results", [])
        if not results:
            print(f"[INFO] No more articles found. Stopping at page {page_num}.")
            break

        # Process each result
        for result_item in results:
            doi_obj = result_item.get("response", {})

            doi = doi_obj.get("doi")
            title = doi_obj.get("title")
            publisher = doi_obj.get("publisher")
            year = doi_obj.get("year")
            oa_status = doi_obj.get("oa_status")
            journal_name = doi_obj.get("journal_name")

            best_oa_loc = doi_obj.get("best_oa_location") or {}
            best_oa_url = best_oa_loc.get("url_for_pdf") or best_oa_loc.get("url")

            # Append row to sheet "All"
            sheet_all.append(
                [doi, title, publisher, year, oa_status, best_oa_url, journal_name]
            )

        print(f"[INFO] Fetched {len(results)} results from page {page_num}.")

        # Be polite: small pause
        time.sleep(wait_s)

    # Save the Excel file
    wb.save(output_xlsx)
    print(f"[DONE] Results saved to '{output_xlsx}'.")


def filter_oa_status_not_closed(excel_file):
    """
    Read the Excel file from the first sheet ("All") and filter out only the
    papers for which 'oa_status' is NOT 'closed'. Write those rows into a
    separate sheet called 'Accessible'.

    Args:
        excel_file (str): Path to the Excel file created by search_unpaywall_by_title.
    """
    # Load existing workbook and get the 'All' sheet
    wb = openpyxl.load_workbook(excel_file)
    sheet_all = wb["All"]

    # Create or replace a sheet for accessible papers
    if "Accessible" in wb.sheetnames:
        del wb["Accessible"]
    sheet_accessible = wb.create_sheet("Accessible")

    # Copy the header row from 'All' to 'Accessible'
    header = [cell.value for cell in sheet_all[1]]
    sheet_accessible.append(header)

    # We know the oa_status column is the 5th in the original structure:
    #   [doi, title, publisher, year, oa_status, best_oa_location_url, journal_name]
    # That means index 4 is 'oa_status'.
    oa_status_index = 4

    # Iterate through the remaining rows in 'All'
    for row in sheet_all.iter_rows(min_row=2, values_only=True):
        if row[oa_status_index] != "closed":
            sheet_accessible.append(row)

    # Save changes
    wb.save(excel_file)
    print(f"[INFO] Filtered data saved in sheet 'Accessible' within '{excel_file}'.")


def download_papers_from_accessible(excel_file, directory="Titian"):
    """
    Reads the 'Accessible' sheet from the specified Excel file and attempts to
    download each paper to a local directory, if a valid URL is present.

    Approach:
      - We look for the 'best_oa_location_url' column and try to download from it.
      - We check whether it is likely a PDF by looking at the file extension
        or the content type from the server.
      - The downloaded file is named using the DOI (slashes replaced by '_').

    Args:
        excel_file (str): The Excel file containing the 'Accessible' sheet.
        directory (str): The directory to store the downloaded papers.
    """
    wb = openpyxl.load_workbook(excel_file)
    if "Accessible" not in wb.sheetnames:
        print("[WARNING] 'Accessible' sheet not found. No downloads will be attempted.")
        return

    sheet_accessible = wb["Accessible"]

    # Ensure the output directory exists
    os.makedirs(directory, exist_ok=True)

    # Identify columns by header name
    # We assume the columns are in the same order as when we wrote them:
    #   [doi, title, publisher, year, oa_status, best_oa_location_url, journal_name]
    # Let's find the indices programmatically:
    headers = [cell.value for cell in sheet_accessible[1]]
    doi_index = headers.index("doi")
    url_index = headers.index("best_oa_location_url")

    # Iterate over the sheet rows, starting from row 2 (since row 1 is header)
    for row in sheet_accessible.iter_rows(min_row=2, values_only=True):
        doi = row[doi_index]
        url = row[url_index]
        if not url:
            print(f"[INFO] No URL for DOI '{doi}', skipping download.")
            continue

        # Create a safe filename from the DOI
        # e.g., 10.1002/anie.201800624 -> 10.1002_anie.201800624
        safe_doi = doi.replace("/", "_")
        local_pdf_path = os.path.join(directory, f"{safe_doi}.pdf")

        try:
            # Try a GET request to see if it's a PDF
            print(f"[INFO] Attempting download from {url}")
            response = requests.get(url, stream=True, timeout=30)

            # Check if this is indeed a PDF by either the content-type or .pdf extension
            content_type = response.headers.get("Content-Type", "").lower()
            is_pdf = content_type.startswith("application/pdf") or url.lower().endswith(
                ".pdf"
            )

            if response.status_code == 200 and is_pdf:
                print(f"[INFO] Downloading PDF for DOI '{doi}' ...")
                with open(local_pdf_path, "wb") as f:
                    for chunk in response.iter_content(chunk_size=8192):
                        if chunk:
                            f.write(chunk)
                print(f"[INFO] Saved: {local_pdf_path}")
            else:
                # Not a direct PDF (e.g., might be an HTML landing page)
                # You could add more logic here to handle special cases.
                print(f"[WARNING] Not a PDF or download failed. URL={url}")
        except requests.exceptions.RequestException as e:
            print(f"[ERROR] Download failed for DOI '{doi}'. Reason: {e}")


if __name__ == "__main__":
    # Example usage
    query_keywords = "Van Gogh paintings"

    # 1. Fetch articles from Unpaywall
    search_unpaywall_by_title(
        keywords=query_keywords,
        email=EMAIL,
        is_oa=None,  # or True/False
        max_pages=3,
        output_xlsx="vg_paintings.xlsx",
        wait_s=1.0,
    )

    # 2. Create an 'Accessible' sheet with only non-closed papers
    filter_oa_status_not_closed("vg_paintings.xlsx")

    # 3. Download the papers from 'Accessible' to directory 'Titian'
    download_papers_from_accessible("vg_paintings.xlsx", directory="Van Gogh")
